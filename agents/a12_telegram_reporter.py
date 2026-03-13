"""
============================================================
AGENT A-12: TELEGRAM REPORTER / COMMAND CENTER — INDUSTRIAL GRADE v5.1
============================================================
The user-facing interface — handles 26 Telegram commands,
morning/evening reports, real-time alerts, inline keyboards,
and full application lifecycle management.

Framework: python-telegram-bot v21
Schedule: 07:15 AM (morning brief) + 10:00 PM (evening summary)

24 Commands:
    /start        — Welcome message + setup wizard
    /help         — Full command reference
    /morning      — Morning brief (top 10, ghost filter, signals)
    /top [N]      — Top N listings by PPO score
    /ocean        — Blue Ocean listings (high prestige, low applicants)
    /internshala  — Live Internshala search
    /dark         — Latest dark channel finds
    /signals      — Active intent signals this week
    /package [id] — Full application package (cover + ATS + intro)
    /ats [id]     — ATS simulation + keyword gap + resume tweaks
    /cover [id]   — 200-word tailored cover letter
    /network [co] — Alumni/warm intro map + outreach draft
    /apply [id]   — Mark listing as applied
    /outcome [id] — Log outcome (interview/reject/offer/ppo)
    /cirs [co]    — Company Intern Readiness Score breakdown
    /research [co]— Full company brief (News+Glassdoor+CIRS)
    /stats        — Weekly funnel + top sector performance
    /health       — Agent heartbeats and system health
    /quota        — Daily API usage (Groq, Cerebras, SerpAPI)
    /export       — Export top listings to formatted text
    /settings     — User preferences (college, specialization)
    /refresh      — Force re-scrape current sources
    /run [agent]  — Run agent NOW (pipeline/scrape/dedup/ghost/enrich/ppo/intent/ats)
    /schedule     — Show full 24-hour schedule + next run times
    /status       — Show currently running background agents
    /cancel       — Cancel a running background agent task
============================================================
"""

import os
import json
import time
import asyncio
import signal
import traceback
import functools
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any

try:
    from loguru import logger
except ImportError:
    import logging
    logger = logging.getLogger(__name__)

try:
    import httpx
    HTTPX_AVAILABLE = True
except ImportError:
    HTTPX_AVAILABLE = False

from core.config import get_config, IST, MBA_CATEGORIES
from core.database import get_db, DatabaseManager, Outcome, OutcomeStatus
from core.ai_router import get_router, AIRouter
from core.security import (
    get_security_manager, SecurityManager,
    require_auth, require_admin,
    ADMIN_TELEGRAM_ID, ADMIN_USERNAME,
)

AGENT_ID = "A-12"
AGENT_NAME = "Telegram Reporter"

# Mini-App URL — set via RENDER_EXTERNAL_URL or MINI_APP_URL env var
# The mini-app should be hosted at this URL (e.g. Cloudflare Pages, Vercel, etc.)
MINI_APP_URL = os.getenv('MINI_APP_URL', '')

# Valid outcome statuses
VALID_OUTCOMES = ['applied', 'shortlisted', 'interview', 'rejected', 'offer', 'ppo', 'withdrawn']

# Message length limit for Telegram
TG_MAX_LEN = 4096


# ============================================================
# ERROR BOUNDARY DECORATOR
# ============================================================

def command_error_boundary(func):
    """
    Industrial-grade error boundary for Telegram command handlers.
    Catches ALL exceptions and sends a user-friendly error message
    instead of silently failing or crashing the bot.
    
    Also enforces authorization — unauthorized users get a denial
    message instead of the command output.
    """
    @functools.wraps(func)
    async def wrapper(self, update, context):
        try:
            # ---- AUTHORIZATION CHECK ----
            # /start is always allowed (for discovery)
            # Admin commands handle their own auth
            cmd_name = func.__name__.replace('_cmd_', '')
            
            # These commands handle their own auth or are always public
            always_allowed = {'start', 'adduser', 'removeuser', 'readduser', 
                            'listusers', 'gencode', 'secstatus',
                            'startpipeline', 'stoppipeline', 'menu'}
            
            if cmd_name not in always_allowed:
                telegram_id = update.effective_user.id if update.effective_user else 0
                if telegram_id:
                    try:
                        sec = get_security_manager()
                        authorized, reason = sec.authorize_command(
                            telegram_id, f'/{cmd_name}'
                        )
                        if not authorized:
                            await update.message.reply_text(
                                f"🔒 Access Denied: {reason}\n\n"
                                f"Your ID: <code>{telegram_id}</code>\n"
                                f"Contact admin to get authorized.",
                                parse_mode='HTML'
                            )
                            return
                    except Exception as auth_err:
                        logger.debug(f"[{AGENT_ID}] Auth check error (allowing): {auth_err}")
            
            return await func(self, update, context)
        except Exception as e:
            cmd_display = func.__name__.replace('_cmd_', '/')
            error_msg = str(e)[:200]
            logger.error(
                f"[{AGENT_ID}] Command {cmd_display} failed: "
                f"{type(e).__name__}: {error_msg}"
            )
            logger.debug(f"[{AGENT_ID}] {cmd_display} traceback:\n{traceback.format_exc()[-500:]}")
            try:
                await update.message.reply_text(
                    f"❌ Error in {cmd_display}: {error_msg}\n\n"
                    f"This has been logged. Try again or use /health to check system status."
                )
            except Exception:
                pass  # Can't even send error message — Telegram might be down
    return wrapper


# ============================================================
# REPORT FORMATTERS
# ============================================================

class ReportFormatter:
    """Formats data for Telegram HTML display with rich, professional formatting."""

    # Source emoji mapping
    SOURCE_EMOJI = {
        'internshala': '🟢', 'naukri': '🔵', 'linkedin': '🟤',
        'greenhouse': '🌿', 'lever': '⚙️', 'indeed': '🟠',
        'iimjobs': '🟡', 'wellfound': '🔶', 'dark_channel': '🌑',
        'glassdoor': '🟩', 'workday': '🔷', 'smartrecruiters': '💼',
        'ashby': '🔘', 'twitter_x': '🐦', 'career_page': '🏢',
        'instahyre': '💎', 'ats_crawler': '🤖',
    }

    # Tier label mapping
    TIER_LABELS = {
        1: 'T1 Elite', 2: 'T2 MNC', 3: 'T3 Unicorn',
        4: 'T4 Startup', 5: 'T5 Niche',
    }

    @classmethod
    def _format_listing_line(cls, i: int, l: Dict, detailed: bool = False) -> str:
        """Format a single listing line with all professional details."""
        title = l.get('title', 'Unknown')
        company = l.get('company', 'Unknown')
        ppo = l.get('ppo_score', 0) or 0
        stipend = l.get('stipend_monthly', 0) or 0
        location = l.get('location', '') or ''
        source = l.get('source', '') or ''
        url = l.get('url', '') or ''
        lid = l.get('id', 0)
        applicants = l.get('applicants', 0) or 0
        duration = l.get('duration_months', 0) or 0
        is_ppo = l.get('is_ppo', False)
        is_wfh = l.get('is_wfh', False)
        is_bo = l.get('is_blue_ocean', False)
        tier = l.get('tier')
        category = l.get('category', '') or ''

        # Tags
        tags = []
        if is_ppo: tags.append('🎯PPO')
        if is_wfh: tags.append('🏠WFH')
        if is_bo: tags.append('🌊Blue Ocean')
        tag_str = ' '.join(tags) if tags else ''

        # Source emoji
        src_emoji = cls.SOURCE_EMOJI.get(source, '📡')

        # Tier label
        tier_label = cls.TIER_LABELS.get(tier, '') if tier else ''

        # Stipend formatting
        if stipend >= 100000:
            stipend_str = f"₹{stipend/100000:.1f}L/mo"
        elif stipend > 0:
            stipend_str = f"₹{stipend:,.0f}/mo"
        else:
            stipend_str = "Unpaid/TBD"

        # Build professional card
        line = f"{'━' * 28}\n"
        line += f"<b>{i}. {title}</b>\n"
        line += f"   🏢 {company}"
        if tier_label:
            line += f" <i>[{tier_label}]</i>"
        line += "\n"

        # Row 2: Location + Duration + Stipend
        row2 = []
        if location:
            row2.append(f"📍{location[:25]}")
        if duration > 0:
            row2.append(f"⏱{duration}mo")
        row2.append(f"💰{stipend_str}")
        line += f"   {' │ '.join(row2)}\n"

        # Row 3: Applicants + PPO Score + Source
        row3 = []
        if applicants > 0:
            if applicants > 500:
                row3.append(f"👥{applicants}⚠️")
            else:
                row3.append(f"👥{applicants}")
        row3.append(f"📊PPO:{ppo:.0f}")
        row3.append(f"{src_emoji}{source}")
        if category:
            row3.append(f"📂{category}")
        line += f"   {' │ '.join(row3)}\n"

        # Row 4: Tags (if any)
        if tag_str:
            line += f"   {tag_str}\n"

        # Row 5: Action buttons
        if url:
            line += f"   🔗 <a href=\"{url}\">APPLY NOW</a> │ /package {lid}\n"
        else:
            line += f"   /package {lid} │ /cover {lid}\n"

        return line

    @classmethod
    def format_jobs_header(cls, total: int, page: int, total_pages: int,
                            sort_by: str, max_duration: int,
                            filters: Dict = None) -> str:
        """Format the header for /jobs command output."""
        sort_labels = {
            'stipend': '💰 Stipend (High→Low)',
            'ppo': '📊 PPO Score',
            'date': '📅 Newest First',
            'duration': '⏱ Duration (Short→Long)',
            'applicants': '👥 Competition (Low→High)',
        }
        sort_label = sort_labels.get(sort_by, sort_by)

        lines = [
            f"📋 <b>MANAGEMENT INTERNSHIPS</b>",
            f"<i>{datetime.now(IST).strftime('%d %b %Y, %I:%M %p IST')}</i>",
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            f"📊 <b>{total}</b> listings │ Page <b>{page}</b>/{total_pages}",
            f"🔀 Sorted by: {sort_label}",
            f"⏱ Duration: ≤ {max_duration} months",
        ]

        if filters:
            filter_parts = []
            if filters.get('category'):
                filter_parts.append(f"📂 {filters['category']}")
            if filters.get('source'):
                filter_parts.append(f"📡 {filters['source']}")
            if filters.get('location'):
                filter_parts.append(f"📍 {filters['location']}")
            if filters.get('min_stipend'):
                filter_parts.append(f"💰 ≥₹{filters['min_stipend']:,.0f}")
            if filter_parts:
                lines.append(f"🔍 Filters: {' │ '.join(filter_parts)}")

        lines.append(f"🚫 Sales/BD/Cold-Calling: <b>AUTO-EXCLUDED</b>")
        lines.append("")
        return '\n'.join(lines)

    @classmethod
    def morning_brief(cls, data: Dict) -> str:
        """Format morning brief report with professional details."""
        total_new = data.get('total_new', 0)
        total_active = data.get('total_active', 0)
        total_raw = data.get('total_raw', 0)
        unprocessed_raw = data.get('unprocessed_raw', 0)
        after_ghost = data.get('after_ghost_filter', 0)
        blue_ocean = data.get('blue_ocean_count', 0)
        signals = data.get('signals_fired', 0)
        top_10 = data.get('top_10', [])
        dark = data.get('dark_finds', [])
        urgent = data.get('urgent_deadlines', [])

        lines = [
            f"🌅 <b>MORNING BRIEF — {datetime.now(IST).strftime('%d %b %Y, %I:%M %p IST')}</b>",
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            f"",
            f"📊 <b>Pipeline Summary</b>",
            f"  New listings (24h): <b>{total_new}</b>",
            f"  After ghost filter: <b>{after_ghost}</b>",
            f"  Total active: <b>{total_active}</b>",
            f"  Blue Ocean alerts: <b>{blue_ocean}</b> 🌊",
            f"  Intent signals: <b>{signals}</b> 📡",
        ]

        # Show pipeline health diagnostics if there are unprocessed items
        if unprocessed_raw > 0:
            lines.append(f"  ⚠️ Unprocessed raw: <b>{unprocessed_raw}</b> (run /run pipeline)")
        if total_raw > 0 and total_active == 0:
            lines.append(f"  📦 Raw scraped: <b>{total_raw}</b> (needs dedup processing)")

        # Source breakdown
        source_data = data.get('source_counts', {})
        if source_data:
            lines.append(f"")
            lines.append(f"📡 <b>Sources</b>")
            for src, cnt in sorted(source_data.items(), key=lambda x: -x[1]):
                if cnt > 0:
                    emoji = cls.SOURCE_EMOJI.get(src, '📡')
                    lines.append(f"  {emoji} {src}: <b>{cnt}</b>")

        lines.append(f"")

        if top_10:
            lines.append(f"🏆 <b>TOP {len(top_10[:10])} BY PPO SCORE</b>")
            lines.append(f"")
            for i, listing in enumerate(top_10[:10], 1):
                lines.append(cls._format_listing_line(i, listing))
        elif total_active == 0 and total_raw > 0:
            lines.append(
                f"📭 {total_raw} raw listings scraped but not yet processed.\n"
                f"Run /run pipeline to dedup, score, and rank them."
            )
        elif total_active == 0:
            lines.append("📭 No listings yet. Run /run pipeline to start scraping.")
        else:
            lines.append(f"📭 No scored listings. {total_active} active listings need PPO scoring.")

        if dark:
            lines.append(f"🌑 <b>Dark Channel:</b> {len(dark)} new finds")

        if urgent:
            lines.append(f"⏰ <b>Urgent (closing soon):</b> {len(urgent)} listings")

        lines.append(f"")
        lines.append(f"💡 /jobs for filtered | /top 25 for more | /ocean for Blue Ocean")
        return '\n'.join(lines)

    @classmethod
    def evening_summary(cls, data: Dict) -> str:
        """Format evening summary report."""
        today_total = data.get('today_total', 0)
        afternoon_new = data.get('afternoon_new', 0)
        applied_today = data.get('applied_today', 0)
        dark_finds = data.get('dark_finds', 0)

        return (
            f"🌆 <b>EVENING SUMMARY — {datetime.now(IST).strftime('%d %b %Y, %I:%M %p IST')}</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"\n"
            f"📊 <b>Today's Numbers</b>\n"
            f"  Total clean listings: <b>{today_total}</b>\n"
            f"  Afternoon new: <b>{afternoon_new}</b>\n"
            f"  Applied today: <b>{applied_today}</b> 📝\n"
            f"  Dark channel finds: <b>{dark_finds}</b>\n"
            f"\n"
            f"💡 /stats for weekly funnel | /health for system status"
        )

    @classmethod
    def listing_detail(cls, listing: Dict) -> str:
        """Format a single listing for detailed view with all information."""
        title = listing.get('title', 'Unknown')
        company = listing.get('company', 'Unknown')
        location = listing.get('location', 'N/A')
        stipend = listing.get('stipend_monthly', 0) or 0
        duration = listing.get('duration_months', 0) or 0
        applicants = listing.get('applicants', 0) or 0
        ppo_score = listing.get('ppo_score', 0)
        ghost_score = listing.get('ghost_score', 0)
        is_ppo = listing.get('is_ppo', False)
        is_wfh = listing.get('is_wfh', False)
        is_bo = listing.get('is_blue_ocean', False)
        source = listing.get('source', 'Unknown')
        url = listing.get('url', '')
        lid = listing.get('id', 0)
        category = listing.get('category', '')
        tier = listing.get('tier')
        sector = listing.get('sector', '')
        description = listing.get('description_text', '')

        # Tags
        tags = []
        if is_ppo: tags.append("🎯 PPO Possible")
        if is_wfh: tags.append("🏠 Work From Home")
        if is_bo: tags.append("🌊 Blue Ocean")
        tag_str = '\n'.join(f"  {t}" for t in tags) if tags else "  None"

        # Tier info
        tier_label = cls.TIER_LABELS.get(tier, 'Unknown') if tier else 'Unknown'
        src_emoji = cls.SOURCE_EMOJI.get(source, '📡')

        # Brief JD snippet
        jd_snippet = ''
        if description:
            clean_desc = description.replace('\n', ' ').strip()[:300]
            jd_snippet = f"\n📝 <b>Description</b>\n<i>{clean_desc}...</i>\n"

        return (
            f"📋 <b>LISTING #{lid} — FULL DETAILS</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"\n"
            f"<b>{title}</b>\n"
            f"🏢 {company} [{tier_label}]\n"
            f"📍 {location}\n"
            f"🏭 Sector: {sector or 'N/A'} | Category: {category or 'N/A'}\n"
            f"\n"
            f"💰 Stipend: {'₹{:,.0f}/month'.format(stipend) if stipend > 0 else 'Not specified'}\n"
            f"⏱ Duration: {f'{duration} months' if duration > 0 else 'Not specified'}\n"
            f"👥 Applicants: {applicants if applicants > 0 else 'Not available'}\n"
            f"\n"
            f"📊 <b>Scores</b>\n"
            f"  PPO Score: <b>{ppo_score:.1f}</b>/100\n"
            f"  Ghost Score: {ghost_score:.0f}/100\n"
            f"\n"
            f"🏷 <b>Tags</b>\n{tag_str}\n"
            f"\n"
            f"📡 Source: {src_emoji} {source}\n"
            f"🔗 {f'<a href=\"{url}\">Open Listing</a>' if url else 'No URL available'}\n"
            f"{jd_snippet}"
            f"\n"
            f"⚡ <b>Actions</b>\n"
            f"  /ats {lid} — ATS simulation\n"
            f"  /cover {lid} — Generate cover letter\n"
            f"  /package {lid} — Full application package\n"
            f"  /apply {lid} — Mark as applied"
        )

    @staticmethod
    def health_report(heartbeats: List[Dict]) -> str:
        """Format agent health report."""
        lines = [
            "💚 <b>Agent Health Dashboard</b>",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "",
        ]

        status_emojis = {
            'idle': '😴', 'running': '🏃', 'error': '❌',
            'completed': '✅', 'disabled': '⛔',
        }

        for h in heartbeats:
            agent_id = h.get('agent_id', '?')
            name = h.get('agent_name', '?')
            status = h.get('status', 'idle')
            emoji = status_emojis.get(status, '❓')
            runs = h.get('total_runs', 0)
            items = h.get('total_items', 0)
            errors = h.get('errors_last_run', 0)
            last_run = h.get('last_run', 'Never')

            lines.append(
                f"{emoji} <b>{agent_id}</b>: {name}\n"
                f"   Runs: {runs} | Items: {items} | "
                f"Errors: {errors}\n"
                f"   Last: {str(last_run)[:19]}"
            )

        return '\n'.join(lines)

    @staticmethod
    def stats_report(stats: Dict) -> str:
        """Format weekly stats report."""
        lines = [
            "📈 <b>Weekly Statistics</b>",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "",
            "<b>Application Funnel:</b>",
        ]

        funnel = stats.get('funnel', {})
        for status, count in funnel.items():
            emoji = {
                'applied': '📝', 'shortlisted': '📋', 'interview': '🎤',
                'rejected': '❌', 'offer': '🎉', 'ppo': '🏆',
            }.get(status, '•')
            lines.append(f"  {emoji} {status.title()}: {count}")

        by_source = stats.get('by_source', {})
        if by_source:
            lines.append(f"\n<b>By Source:</b>")
            for src, count in sorted(by_source.items(), key=lambda x: -x[1]):
                lines.append(f"  {src}: {count}")

        by_tier = stats.get('by_tier', {})
        if by_tier:
            lines.append(f"\n<b>By Company Tier:</b>")
            for tier, count in sorted(by_tier.items()):
                lines.append(f"  {tier}: {count}")

        return '\n'.join(lines)


# ============================================================
# TELEGRAM BOT
# ============================================================

# ============================================================
# CONSTANTS
# ============================================================

# Max retries for bot startup when Conflict is hit
BOT_START_MAX_RETRIES = 5
BOT_START_RETRY_BASE_DELAY = 15  # seconds — must be long enough for old instance to die
BOT_CONFLICT_COOLDOWN = 20  # seconds to wait after killing stale session

# Pre-flight drain: number of rapid getUpdates calls to forcefully
# terminate any lingering long-poll connection from a previous instance.
PREFLIGHT_DRAIN_ROUNDS = 8
PREFLIGHT_DRAIN_INTERVAL = 2  # seconds between drain calls

# Pre-flight: max time to wait for old instance to die (checked via health endpoint)
PREFLIGHT_OLD_INSTANCE_CHECK_TIMEOUT = 30  # seconds
PREFLIGHT_OLD_INSTANCE_CHECK_INTERVAL = 5  # seconds

# Instance lock file — prevents concurrent polling on same machine
INSTANCE_LOCK_FILE = "data/.telegram_polling.lock"


class TelegramReporter:
    """
    Telegram bot command center with 26 commands.
    Uses python-telegram-bot v21 with async handlers.

    ROBUSTNESS FEATURES:
        1. Pre-flight cleanup: deletes any stale webhook AND
           calls getUpdates with offset=-1 to flush the queue
           and release any lingering polling lock.
        2. Conflict-aware startup with exponential backoff:
           If Conflict is detected, waits and retries up to 5 times.
        3. Custom error handler that catches Conflict during runtime
           and auto-restarts polling instead of crashing.
        4. Graceful shutdown ensures polling is fully stopped
           before the process exits (SIGTERM from Render).
    """

    def __init__(self):
        self.db = get_db()
        self.config = get_config()
        self.router = get_router()
        self.formatter = ReportFormatter()
        self.security = get_security_manager()
        self._app = None
        self._running = False
        self._restart_lock = asyncio.Lock()
        self._restart_count = 0
        self._max_runtime_restarts = 10  # max auto-restarts during runtime

        # Background task tracking for /run, /status, /cancel
        self._running_tasks: Dict[str, Dict[str, Any]] = {}
        self._task_lock = asyncio.Lock()
        
        # Instance lock for single-instance guarantee
        self._lock_fd = None
        
        # Pipeline control flag for admin stop
        self._pipeline_stop_requested = False

    # ================================================================
    # PRE-FLIGHT: CLEAN STALE SESSIONS
    # ================================================================

    async def _wait_for_old_instance_death(self):
        """
        If running on Render, check the external URL to see if the
        OLD instance has stopped its Telegram polling. This is more
        reliable than a fixed sleep timer.
        """
        import aiohttp

        external_url = os.getenv('RENDER_EXTERNAL_URL', '')
        if not external_url:
            return  # Not on Render or no external URL

        status_url = f"{external_url}/telegram-status"
        timeout = aiohttp.ClientTimeout(total=10)

        logger.info(
            f"[{AGENT_ID}] Checking old instance Telegram status at "
            f"{status_url}..."
        )

        elapsed = 0
        while elapsed < PREFLIGHT_OLD_INSTANCE_CHECK_TIMEOUT:
            try:
                async with aiohttp.ClientSession(timeout=timeout) as session:
                    async with session.get(status_url) as resp:
                        if resp.status == 200:
                            import json as _json
                            data = _json.loads(await resp.text())
                            tg_running = data.get('telegram_running', False)
                            if not tg_running:
                                logger.info(
                                    f"[{AGENT_ID}] Old instance confirms "
                                    f"Telegram is OFF — safe to proceed"
                                )
                                return
                            else:
                                logger.info(
                                    f"[{AGENT_ID}] Old instance still has "
                                    f"Telegram ON — waiting..."
                                )
                        else:
                            # Non-200 means old instance may be dead
                            logger.info(
                                f"[{AGENT_ID}] Old instance returned "
                                f"HTTP {resp.status} — likely shutting down"
                            )
                            return
            except Exception as e:
                # Connection error = old instance is dead
                logger.info(
                    f"[{AGENT_ID}] Old instance unreachable ({e}) — "
                    f"proceeding"
                )
                return

            await asyncio.sleep(PREFLIGHT_OLD_INSTANCE_CHECK_INTERVAL)
            elapsed += PREFLIGHT_OLD_INSTANCE_CHECK_INTERVAL

        logger.warning(
            f"[{AGENT_ID}] Old instance check timed out after "
            f"{PREFLIGHT_OLD_INSTANCE_CHECK_TIMEOUT}s — proceeding anyway"
        )

    async def _kill_stale_sessions(self, token: str):
        """
        NUCLEAR CLEANUP — Call this BEFORE starting polling.

        The REAL fix for Conflict errors:

        1. Call Telegram Bot API `close()` — this explicitly tells
           Telegram's servers to LOG OUT any active session for this
           bot token. Unlike getUpdates(offset=-1) which only interrupts
           the current long-poll, `close` server-side revokes the session.

        2. Call `logOut()` as a fallback — forces re-authentication.
           (Only needed in extreme cases, skipped if close succeeds.)

        3. deleteWebhook(drop_pending_updates=True) — standard cleanup.

        4. DRAIN LOOP — multiple rapid getUpdates calls to flush any
           queued updates and interrupt PTB's auto-retry on old instance.

        5. Final getUpdates with offset to acknowledge all pending.

        6. Cooldown for Telegram servers to release connection state.
        """
        import aiohttp

        base = f"https://api.telegram.org/bot{token}"
        timeout = aiohttp.ClientTimeout(total=15)

        try:
            async with aiohttp.ClientSession(timeout=timeout) as session:
                # Step 1: CLOSE the bot session server-side
                # This is the KEY fix — `close` tells Telegram to drop
                # ALL active connections for this bot token.
                logger.info(f"[{AGENT_ID}] Pre-flight: calling /close...")
                try:
                    async with session.post(f"{base}/close") as resp:
                        result = await resp.json()
                        logger.info(
                            f"[{AGENT_ID}] Pre-flight: /close → "
                            f"{result.get('ok', False)} "
                            f"(desc: {result.get('description', 'n/a')})"
                        )
                except Exception as close_err:
                    logger.warning(
                        f"[{AGENT_ID}] Pre-flight: /close failed: {close_err}"
                    )

                # After /close, Telegram requires a 10-second cooldown
                # before the bot can make new requests
                logger.info(
                    f"[{AGENT_ID}] Pre-flight: post-close cooldown (12s)..."
                )
                await asyncio.sleep(12)

                # Step 2: Delete webhook + flush pending
                logger.info(f"[{AGENT_ID}] Pre-flight: deleting webhook...")
                try:
                    async with session.post(
                        f"{base}/deleteWebhook",
                        json={"drop_pending_updates": True}
                    ) as resp:
                        result = await resp.json()
                        logger.info(
                            f"[{AGENT_ID}] Pre-flight: deleteWebhook → "
                            f"{result.get('ok', False)}"
                        )
                except Exception as wh_err:
                    logger.warning(
                        f"[{AGENT_ID}] Pre-flight: deleteWebhook failed: "
                        f"{wh_err}"
                    )

                # Step 3: DRAIN LOOP — repeatedly call getUpdates to
                # flush any remaining queued updates and interrupt
                # the old instance's retry mechanism.
                logger.info(
                    f"[{AGENT_ID}] Pre-flight: drain loop "
                    f"({PREFLIGHT_DRAIN_ROUNDS} rounds, "
                    f"{PREFLIGHT_DRAIN_INTERVAL}s apart)..."
                )
                last_update_id = None
                conflict_count = 0
                for i in range(1, PREFLIGHT_DRAIN_ROUNDS + 1):
                    try:
                        async with session.post(
                            f"{base}/getUpdates",
                            json={"offset": -1, "timeout": 1}
                        ) as resp:
                            result = await resp.json()
                            if result.get('ok'):
                                updates = result.get('result', [])
                                if updates:
                                    last_update_id = updates[-1].get(
                                        'update_id', 0
                                    )
                                logger.debug(
                                    f"[{AGENT_ID}] Pre-flight drain #{i}: "
                                    f"ok=True, updates={len(updates)}"
                                )
                            else:
                                err_code = result.get('error_code', 0)
                                desc = result.get('description', '')
                                if err_code == 409:  # Conflict
                                    conflict_count += 1
                                    logger.info(
                                        f"[{AGENT_ID}] Pre-flight drain #{i}: "
                                        f"Conflict detected (old instance alive)"
                                    )
                                else:
                                    logger.debug(
                                        f"[{AGENT_ID}] Pre-flight drain #{i}: "
                                        f"err={err_code} desc={desc}"
                                    )
                    except Exception as drain_err:
                        logger.debug(
                            f"[{AGENT_ID}] Pre-flight drain #{i} error: "
                            f"{drain_err}"
                        )
                    if i < PREFLIGHT_DRAIN_ROUNDS:
                        await asyncio.sleep(PREFLIGHT_DRAIN_INTERVAL)

                logger.info(
                    f"[{AGENT_ID}] Pre-flight: drain loop complete "
                    f"({PREFLIGHT_DRAIN_ROUNDS} rounds, "
                    f"{conflict_count} conflicts)"
                )

                # Step 4: Acknowledge the last update so the queue is clean
                if last_update_id is not None:
                    try:
                        async with session.post(
                            f"{base}/getUpdates",
                            json={
                                "offset": last_update_id + 1,
                                "timeout": 1,
                            }
                        ) as resp:
                            await resp.json()
                            logger.info(
                                f"[{AGENT_ID}] Pre-flight: acknowledged up to "
                                f"update_id {last_update_id}"
                            )
                    except Exception:
                        pass

                # Step 5: Final cooldown
                cooldown = 8 if conflict_count > 0 else 3
                logger.info(
                    f"[{AGENT_ID}] Pre-flight: final cooldown ({cooldown}s)..."
                )
                await asyncio.sleep(cooldown)

        except Exception as e:
            logger.warning(
                f"[{AGENT_ID}] Pre-flight cleanup error (non-fatal): {e}"
            )
            # Non-fatal — we still try to start

    # ================================================================
    # BOT START WITH RETRY
    # ================================================================

    def _build_app(self, token: str):
        """
        Build a fresh TGApplication with all handlers registered.
        Extracted to avoid duplication in retry loop.
        """
        from telegram.ext import (
            Application as TGApplication,
            CommandHandler,
        )

        app = (
            TGApplication.builder()
            .token(token)
            .connect_timeout(30)
            .read_timeout(30)
            .write_timeout(30)
            .pool_timeout(15)
            .get_updates_connect_timeout(20)
            .get_updates_read_timeout(45)
            .get_updates_write_timeout(20)
            .get_updates_pool_timeout(15)
            .build()
        )

        # Register custom error handler for runtime Conflict errors
        app.add_error_handler(self._on_telegram_error)

        # Register all command handlers (26 user + 6 admin + pipeline admin)
        commands = {
            'start': self._cmd_start,
            'help': self._cmd_help,
            'morning': self._cmd_morning,
            'top': self._cmd_top,
            'ocean': self._cmd_ocean,
            'internshala': self._cmd_internshala,
            'dark': self._cmd_dark,
            'signals': self._cmd_signals,
            'package': self._cmd_package,
            'ats': self._cmd_ats,
            'cover': self._cmd_cover,
            'network': self._cmd_network,
            'apply': self._cmd_apply,
            'outcome': self._cmd_outcome,
            'cirs': self._cmd_cirs,
            'research': self._cmd_research,
            'stats': self._cmd_stats,
            'health': self._cmd_health,
            'quota': self._cmd_quota,
            'export': self._cmd_export,
            'settings': self._cmd_settings,
            'refresh': self._cmd_refresh,
            'run': self._cmd_run,
            'schedule': self._cmd_schedule,
            'status': self._cmd_status,
            'cancel': self._cmd_cancel,
            'queue': self._cmd_queue,
            'autoapply': self._cmd_autoapply,
            'appstatus': self._cmd_appstatus,
            'loadall': self._cmd_loadall,
            'filter': self._cmd_filter,
            'sources': self._cmd_sources,
            'browse': self._cmd_browse,
            'cfstatus': self._cmd_cfstatus,
            'reprocess': self._cmd_reprocess,
            'jobs': self._cmd_jobs,
            # Security & Admin commands
            'adduser': self._cmd_adduser,
            'removeuser': self._cmd_removeuser,
            'readduser': self._cmd_readduser,
            'listusers': self._cmd_listusers,
            'gencode': self._cmd_gencode,
            'secstatus': self._cmd_secstatus,
            # Admin pipeline control
            'startpipeline': self._cmd_startpipeline,
            'stoppipeline': self._cmd_stoppipeline,
            # Mini-App access
            'miniapp': self._cmd_miniapp,
            'webapp': self._cmd_miniapp,  # alias
            # Menu trigger
            'menu': self._cmd_menu,
        }

        for cmd_name, handler_fn in commands.items():
            app.add_handler(CommandHandler(cmd_name, handler_fn))

        # Add text message handler for menu button presses
        from telegram.ext import MessageHandler, filters as tg_filters
        app.add_handler(MessageHandler(
            tg_filters.TEXT & ~tg_filters.COMMAND,
            self._handle_menu_button_press
        ))

        return app

    async def _safe_cleanup_app(self, app):
        """
        Safely cleanup a TGApplication instance.
        Each step is individually try/excepted and timed out.
        """
        if not app:
            return

        # Stop updater
        try:
            if app.updater and app.updater.running:
                await asyncio.wait_for(app.updater.stop(), timeout=5.0)
        except Exception:
            pass

        # Stop application
        try:
            if app.running:
                await asyncio.wait_for(app.stop(), timeout=5.0)
        except Exception:
            pass

        # Shutdown
        try:
            await asyncio.wait_for(app.shutdown(), timeout=5.0)
        except Exception:
            pass

    def _acquire_instance_lock(self) -> bool:
        """
        Acquire a file-based instance lock to prevent multiple
        bot instances on the same machine from polling simultaneously.
        Uses fcntl (POSIX) for atomic locking.
        """
        import fcntl
        lock_path = INSTANCE_LOCK_FILE
        os.makedirs(os.path.dirname(lock_path), exist_ok=True)
        try:
            self._lock_fd = open(lock_path, 'w')
            fcntl.flock(self._lock_fd.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            self._lock_fd.write(f"{os.getpid()}\n")
            self._lock_fd.flush()
            logger.info(f"[{AGENT_ID}] Instance lock acquired (PID {os.getpid()})")
            return True
        except (IOError, OSError):
            logger.warning(
                f"[{AGENT_ID}] Another instance holds the polling lock. "
                f"This instance will NOT start polling."
            )
            if self._lock_fd:
                self._lock_fd.close()
                self._lock_fd = None
            return False
    
    def _release_instance_lock(self):
        """Release the file-based instance lock."""
        import fcntl
        if self._lock_fd:
            try:
                fcntl.flock(self._lock_fd.fileno(), fcntl.LOCK_UN)
                self._lock_fd.close()
            except Exception:
                pass
            self._lock_fd = None
            logger.info(f"[{AGENT_ID}] Instance lock released")

    async def start_bot(self):
        """
        Initialize and start the Telegram bot with full robustness:
        0. Acquire instance lock (prevents dual polling on same machine)
        1. Wait for old instance to die (Render health check)
        2. Kill stale sessions via Telegram /close API
        3. Build application with error handler
        4. Start polling with exponential backoff retries
        """
        token = self.config.telegram.bot_token
        if not token:
            logger.error(f"[{AGENT_ID}] TG_BOT_TOKEN not set!")
            return

        try:
            from telegram import Update, Bot
            from telegram.ext import (
                Application as TGApplication,
                CommandHandler, ContextTypes,
                MessageHandler, filters,
            )
            from telegram.error import Conflict, TimedOut, NetworkError
        except ImportError:
            logger.error(f"[{AGENT_ID}] python-telegram-bot not installed")
            return

        # ---- STEP -1: Acquire instance lock ----
        if not self._acquire_instance_lock():
            logger.error(
                f"[{AGENT_ID}] SKIPPING bot start — another instance is polling. "
                f"This prevents the getUpdates Conflict error."
            )
            return

        # ---- STEP 0: Check if old instance is still alive ----
        await self._wait_for_old_instance_death()

        # ---- STEP 1: Kill any stale polling session ----
        await self._kill_stale_sessions(token)

        # ---- STEP 2: Build application ----
        self._app = self._build_app(token)

        logger.info(
            f"[{AGENT_ID}] Bot starting with 26 commands "
            f"(max {BOT_START_MAX_RETRIES} attempts)..."
        )

        # ---- STEP 3: Start polling with retry ----
        last_error = None
        for attempt in range(1, BOT_START_MAX_RETRIES + 1):
            try:
                await self._app.initialize()
                await self._app.start()
                await self._app.updater.start_polling(
                    drop_pending_updates=True,
                    allowed_updates=["message", "callback_query"],
                    poll_interval=1.0,
                )
                self._running = True
                logger.info(
                    f"[{AGENT_ID}] Telegram bot is running! "
                    f"(started on attempt {attempt}/{BOT_START_MAX_RETRIES})"
                )
                
                # Set up bot commands menu and Open App button
                await self._setup_bot_menu()
                
                return  # SUCCESS

            except Conflict as e:
                last_error = e
                # Calculate delay with capped exponential backoff
                delay = min(
                    BOT_START_RETRY_BASE_DELAY * (2 ** (attempt - 1)),
                    120  # cap at 2 minutes
                )
                logger.warning(
                    f"[{AGENT_ID}] Conflict on attempt "
                    f"{attempt}/{BOT_START_MAX_RETRIES}: {e}. "
                    f"Retrying in {delay}s..."
                )

                # Clean up current app
                await self._safe_cleanup_app(self._app)

                # Wait before retrying
                await asyncio.sleep(delay)

                # Re-kill stale sessions (use /close again)
                await self._kill_stale_sessions(token)

                # Rebuild the application for retry
                self._app = self._build_app(token)

            except Exception as e:
                last_error = e
                logger.error(
                    f"[{AGENT_ID}] Bot start attempt {attempt} failed: "
                    f"{type(e).__name__}: {e}"
                )

                # Clean up
                await self._safe_cleanup_app(self._app)

                if attempt < BOT_START_MAX_RETRIES:
                    delay = min(
                        BOT_START_RETRY_BASE_DELAY * attempt,
                        90
                    )
                    logger.info(
                        f"[{AGENT_ID}] Waiting {delay}s before retry..."
                    )
                    await asyncio.sleep(delay)

                    # Rebuild for retry
                    self._app = self._build_app(token)

        # All retries exhausted
        logger.error(
            f"[{AGENT_ID}] Bot failed to start after "
            f"{BOT_START_MAX_RETRIES} attempts. "
            f"Last error: {last_error}"
        )
        logger.error(
            f"[{AGENT_ID}] System will continue WITHOUT Telegram. "
            f"Fix: ensure only ONE instance uses this bot token."
        )

    # ================================================================
    # RUNTIME ERROR HANDLER (catches Conflict during polling)
    # ================================================================

    async def _on_telegram_error(self, update, context):
        """
        Custom error handler for runtime errors.

        CRITICAL CHANGE: We do NOT auto-restart on Conflict errors.

        WHY: Auto-restarting polling on Conflict creates a feedback loop:
          1. Instance A is polling
          2. Instance B starts polling → A gets Conflict
          3. A's error handler restarts polling → B gets Conflict
          4. B's error handler restarts polling → A gets Conflict
          5. → infinite loop of Conflict errors

        CORRECT BEHAVIOR: If we get a Conflict, it means another instance
        legitimately took over. We should STOP polling and let it win.
        The old instance will be killed by SIGTERM shortly anyway.
        """
        from telegram.error import Conflict, TimedOut, NetworkError

        error = context.error

        if isinstance(error, Conflict):
            logger.warning(
                f"[{AGENT_ID}] RUNTIME Conflict: another instance took over. "
                f"STOPPING polling (this is expected during redeployment)."
            )
            # Do NOT restart. Let the other instance win.
            # PTB will keep retrying internally — we stop it.
            asyncio.create_task(self._surrender_polling())

        elif isinstance(error, (TimedOut, NetworkError)):
            logger.warning(
                f"[{AGENT_ID}] Network issue (auto-recoverable): {error}"
            )
            # PTB handles these automatically, just log

        else:
            logger.error(
                f"[{AGENT_ID}] Unhandled error: {type(error).__name__}: {error}"
            )

    async def _surrender_polling(self):
        """
        Gracefully surrender polling when a Conflict is detected.
        This means another instance legitimately started polling.
        We stop our updater so we don't keep fighting.
        """
        async with self._restart_lock:
            try:
                if self._app and self._app.updater and self._app.updater.running:
                    logger.info(
                        f"[{AGENT_ID}] Surrendering: stopping updater..."
                    )
                    await asyncio.wait_for(
                        self._app.updater.stop(), timeout=10.0
                    )
                    self._running = False
                    logger.info(
                        f"[{AGENT_ID}] Updater stopped. "
                        f"This instance will NOT restart polling."
                    )
            except Exception as e:
                logger.error(
                    f"[{AGENT_ID}] Surrender stop error: {e}"
                )

    # ================================================================
    # GRACEFUL SHUTDOWN
    # ================================================================

    async def stop_bot(self):
        """
        Stop the Telegram bot with FULL cleanup.

        Order matters:
        1. Stop polling (releases getUpdates lock)
        2. Call Telegram /close API to release server-side session
        3. Stop application
        4. Shutdown (cleanup resources)

        The /close call is CRITICAL for Render redeployments:
        it tells Telegram's servers to drop the connection immediately,
        so the NEW instance can start polling without conflict.
        """
        if not self._app:
            return

        logger.info(f"[{AGENT_ID}] Stopping bot (4-step shutdown)...")

        # Step 1: Stop updater/polling FIRST
        try:
            if self._app.updater and self._app.updater.running:
                await asyncio.wait_for(
                    self._app.updater.stop(), timeout=8.0
                )
                logger.info(f"[{AGENT_ID}] Updater stopped")
        except asyncio.TimeoutError:
            logger.warning(f"[{AGENT_ID}] Updater stop timed out (8s)")
        except Exception as e:
            logger.error(f"[{AGENT_ID}] Updater stop error: {e}")

        # Step 2: Call /close to release server-side session
        # This is the KEY step that prevents the next instance from
        # hitting Conflict. It tells Telegram's backend to drop
        # our session immediately.
        try:
            import aiohttp
            token = self.config.telegram.bot_token
            if token:
                base = f"https://api.telegram.org/bot{token}"
                timeout = aiohttp.ClientTimeout(total=5)
                async with aiohttp.ClientSession(timeout=timeout) as session:
                    async with session.post(f"{base}/close") as resp:
                        result = await resp.json()
                        logger.info(
                            f"[{AGENT_ID}] /close called: "
                            f"{result.get('ok', False)}"
                        )
        except Exception as e:
            logger.warning(
                f"[{AGENT_ID}] /close call failed (non-critical): {e}"
            )

        # Step 3: Stop the application
        try:
            if self._app.running:
                await asyncio.wait_for(
                    self._app.stop(), timeout=8.0
                )
                logger.info(f"[{AGENT_ID}] Application stopped")
        except asyncio.TimeoutError:
            logger.warning(f"[{AGENT_ID}] App stop timed out (8s)")
        except Exception as e:
            logger.error(f"[{AGENT_ID}] App stop error: {e}")

        # Step 4: Shutdown (cleanup)
        try:
            await asyncio.wait_for(
                self._app.shutdown(), timeout=8.0
            )
            logger.info(f"[{AGENT_ID}] Application shutdown complete")
        except asyncio.TimeoutError:
            logger.warning(f"[{AGENT_ID}] App shutdown timed out (8s)")
        except Exception as e:
            logger.error(f"[{AGENT_ID}] App shutdown error: {e}")

        self._running = False
        
        # Release instance lock
        self._release_instance_lock()
        
        logger.info(f"[{AGENT_ID}] Bot fully stopped (session + lock released)")

    async def send_message(self, text: str, chat_id: str = None):
        """Send a message to configured chat, with auto-splitting on newlines."""
        if chat_id is None:
            chat_id = self.config.telegram.chat_id
        if not chat_id:
            logger.warning(f"[{AGENT_ID}] No chat_id configured")
            return

        try:
            from telegram import Bot
            bot = Bot(token=self.config.telegram.bot_token)

            # Smart split on newlines for long messages
            chunks = []
            remaining = text
            while remaining:
                if len(remaining) <= TG_MAX_LEN:
                    chunks.append(remaining)
                    break
                split_pos = remaining.rfind('\n', 0, TG_MAX_LEN)
                if split_pos == -1 or split_pos < TG_MAX_LEN // 2:
                    split_pos = TG_MAX_LEN
                chunks.append(remaining[:split_pos])
                remaining = remaining[split_pos:].lstrip('\n')

            for chunk in chunks:
                if not chunk.strip():
                    continue
                try:
                    await bot.send_message(
                        chat_id=chat_id, text=chunk, parse_mode='HTML'
                    )
                except Exception:
                    # Fallback without HTML parse mode
                    try:
                        await bot.send_message(chat_id=chat_id, text=chunk)
                    except Exception as e2:
                        logger.error(f"[{AGENT_ID}] Send fallback also failed: {e2}")
        except Exception as e:
            logger.error(f"[{AGENT_ID}] Send message failed: {e}")

    # ================================================================
    # COMMAND HANDLERS — 22 Commands
    # ================================================================

    @command_error_boundary
    async def _cmd_start(self, update, context):
        """Welcome message and setup wizard. Always accessible (for auth flow)."""
        telegram_id = update.effective_user.id if update.effective_user else 0
        is_authorized = self.security.is_authorized(telegram_id)
        is_admin = self.security.is_admin(telegram_id)
        
        if not is_authorized:
            msg = (
                "🔒 <b>Operation First Mover v5.3</b>\n"
                "━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "This bot requires authorization.\n"
                "Please contact the admin to get access.\n\n"
                f"Your Telegram ID: <code>{telegram_id}</code>\n"
                "Share this with the admin."
            )
            await update.message.reply_text(msg, parse_mode='HTML')
            return
        
        admin_section = ""
        if is_admin:
            admin_section = (
                "\n\n🔐 <b>Admin Commands:</b>\n"
                "/adduser — Add authorized user\n"
                "/removeuser — Remove user\n"
                "/listusers — List all users\n"
                "/secstatus — Security dashboard"
            )
        
        # Get user's access code for mini-app
        user_data = self.security.get_user(telegram_id)
        access_code = user_data.get('access_code', '') if user_data else ''
        
        miniapp_section = ""
        if MINI_APP_URL:
            miniapp_section = (
                "\n\n📱 <b>InternHub Pro (Mini App):</b>\n"
                "/miniapp — Open the Mini App\n"
            )
            if access_code:
                miniapp_section += (
                    f"🔑 Your access code: <code>{access_code}</code>\n"
                )
        
        msg = (
            "⚡ <b>Operation First Mover v5.3</b>\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            "Your zero-cost MBA internship hunting agent.\n\n"
            "🤖 <b>12 AI agents</b> working 24/7\n"
            "📊 <b>1080+</b> Indian companies tracked\n"
            "🔍 <b>8+</b> job boards scraped daily\n"
            "💰 Total cost: <b>₹0.00/day</b>\n"
            "🔐 <b>Security:</b> Enabled\n\n"
            "🎯 <b>Start Here:</b>\n"
            "/jobs — Browse filtered internships (no sales!)\n"
            "/morning — Full morning brief\n"
            "/top 20 — Top 20 by PPO score\n\n"
            "🚀 <b>Run agents on demand:</b>\n"
            "/run pipeline — Full scrape+process+report\n"
            "/run scrape — Just scrape now\n\n"
            f"Type /help for all commands.{miniapp_section}{admin_section}"
        )
        
        # Send with inline keyboard for Mini App if URL is configured
        # ALSO always send persistent reply keyboard
        keyboard_reply = self._get_main_reply_keyboard(is_admin=is_admin)
        
        if MINI_APP_URL:
            try:
                from telegram import InlineKeyboardButton, InlineKeyboardMarkup, WebAppInfo
                inline_keyboard = InlineKeyboardMarkup([
                    [InlineKeyboardButton(
                        "📱 Open InternHub Pro",
                        web_app=WebAppInfo(url=MINI_APP_URL)
                    )],
                    [InlineKeyboardButton(
                        "📋 Browse Jobs", callback_data="cmd_jobs"
                    ), InlineKeyboardButton(
                        "📊 Morning Brief", callback_data="cmd_morning"
                    )],
                ])
                # Send first message with inline keyboard
                await update.message.reply_text(
                    msg, parse_mode='HTML', reply_markup=inline_keyboard
                )
                # Then send a follow-up with the persistent reply keyboard
                await update.message.reply_text(
                    "👇 <b>Quick access menu below!</b>\nTap any button for instant commands.",
                    parse_mode='HTML',
                    reply_markup=keyboard_reply,
                )
            except Exception as e:
                logger.debug(f"[{AGENT_ID}] Inline keyboard failed, sending plain: {e}")
                await update.message.reply_text(
                    msg, parse_mode='HTML', reply_markup=keyboard_reply
                )
        else:
            await update.message.reply_text(
                msg, parse_mode='HTML', reply_markup=keyboard_reply
            )

    @command_error_boundary
    async def _cmd_help(self, update, context):
        """Full command reference."""
        # Auth check
        telegram_id = update.effective_user.id if update.effective_user else 0
        authorized, reason = self.security.authorize_command(telegram_id, '/help')
        if not authorized:
            await update.message.reply_text(f"🔒 {reason}")
            return
        
        is_admin = self.security.is_admin(telegram_id)
        
        msg = (
            "📖 <b>Command Reference (38+ Commands)</b>\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "📋 <b>Browse & Discover</b>\n"
            "/jobs [sort] [page] — <b>SMART filtered (no sales!)</b>\n"
            "  Sorts: stipend, ppo, date, duration, applicants\n"
            "  Filters: /jobs marketing, /jobs mumbai, /jobs 10k\n"
            "/loadall [page] [cat] [src] — Browse ALL listings\n"
            "/browse [category] — Category browser\n"
            "/filter — Available filters & counts\n"
            "/sources — Source health dashboard\n"
            "/top [N] — Top N by PPO score (default 10)\n"
            "/ocean — Blue Ocean listings\n"
            "/export [N] — Export to Excel\n\n"
            "📊 <b>Reports</b>\n"
            "/morning — Full morning brief\n"
            "/dark — Dark channel finds\n"
            "/signals — Active intent signals\n"
            "/stats — Weekly funnel stats\n\n"
            "🔍 <b>Search</b>\n"
            "/internshala [query] — Live search\n"
            "/refresh — Force re-scrape all sources\n\n"
            "📝 <b>Application</b>\n"
            "/package [id] — Full app package\n"
            "/ats [id] — ATS keyword simulation\n"
            "/cover [id] — AI cover letter\n"
            "/network [company] — Alumni map\n"
            "/apply [id] — Mark as applied\n"
            "/outcome [id] [result] — Log result\n\n"
            "🤖 <b>Auto-Apply (A-13)</b>\n"
            "/queue [id] — Add to auto-apply queue\n"
            "/queue top [N] — Queue top N listings\n"
            "/autoapply [N] — Run auto-apply (max N)\n"
            "/appstatus — Application history\n\n"
            "🏢 <b>Company Intel</b>\n"
            "/cirs [company] — CIRS breakdown\n"
            "/research [company] — Full research\n\n"
            "🚀 <b>Agent Control</b>\n"
            "/run — Run agents NOW (see /run for options)\n"
            "/schedule — Full 24h schedule\n"
            "/status — Running tasks\n"
            "/cancel [task] — Cancel task\n\n"
            "⚙️ <b>System</b>\n"
            "/health — Agent heartbeats\n"
            "/quota — API usage\n"
            "/cfstatus — Cloudflare /crawl status\n"
            "/reprocess — Raw listing status/reset\n"
            "/settings — Preferences\n\n"
            "📱 <b>Mini App</b>\n"
            "/miniapp — Open InternHub Pro mini app\n"
            "/webapp — Alias for /miniapp"
        )
        
        if is_admin:
            msg += (
                "\n\n🔐 <b>Admin (Your Eyes Only)</b>\n"
                "/adduser <user> <id> — Authorize user\n"
                "/removeuser <id> — Deactivate user\n"
                "/readduser <id> — Re-enable user\n"
                "/listusers — All authorized users\n"
                "/gencode <id> — Regenerate access code\n"
                "/secstatus — Security dashboard\n\n"
                "🚀 <b>Pipeline Control (Admin)</b>\n"
                "/startpipeline — Start full pipeline\n"
                "/stoppipeline — Stop running pipeline\n"
                "/run <agent> — Run individual agent"
            )
        
        await update.message.reply_text(msg, parse_mode='HTML')

    @command_error_boundary
    async def _cmd_morning(self, update, context):
        """Morning brief report."""
        await update.message.reply_text("🌅 Generating morning brief...")
        data = self.db.get_morning_brief_data()
        msg = self.formatter.morning_brief(data)
        await self._send_long_message(update, msg)

    @command_error_boundary
    async def _cmd_top(self, update, context):
        """Top N listings by PPO score with professional formatting."""
        n = 10
        if context.args:
            try:
                n = int(context.args[0])
                n = max(1, min(50, n))
            except ValueError:
                pass

        listings = self.db.get_top_listings(n=n)
        if not listings:
            await update.message.reply_text(
                "📊 No listings available yet.\n"
                "Run /run pipeline to scrape and process listings."
            )
            return

        lines = [
            f"🏆 <b>Top {len(listings)} by PPO Score</b>",
            f"<i>{datetime.now(IST).strftime('%d %b %Y, %I:%M %p IST')}</i>",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", "",
        ]

        for i, l in enumerate(listings, 1):
            lines.append(self.formatter._format_listing_line(i, l))

        lines.append(f"💡 /jobs for filtered | /export {n} for Excel | /ocean for Blue Ocean")
        await self._send_long_message(update, '\n'.join(lines))

    @command_error_boundary
    async def _cmd_jobs(self, update, context):
        """
        Super-rich management internship browser with smart filtering.

        Usage:
            /jobs                — Page 1, sorted by stipend, ≤3mo, no sales
            /jobs 2              — Page 2
            /jobs stipend 2      — Sort by stipend, page 2
            /jobs ppo            — Sort by PPO score
            /jobs date           — Newest first
            /jobs duration       — Shortest first
            /jobs applicants     — Least competitive first
            /jobs marketing      — Filter by category
            /jobs linkedin       — Filter by source
            /jobs mumbai         — Filter by location
            /jobs 5mo            — Duration ≤ 5 months
            /jobs 10k            — Min stipend ₹10,000

        All results ALWAYS exclude sales/BD/cold-calling roles.
        """
        import re as _re

        # Defaults
        page = 1
        sort_by = 'stipend'
        max_duration = 3
        category_filter = None
        source_filter = None
        location_filter = None
        min_stipend = 0
        per_page = 15

        valid_sorts = {'stipend', 'ppo', 'date', 'duration', 'applicants'}
        valid_sources = {
            'internshala', 'linkedin', 'naukri', 'indeed', 'iimjobs',
            'greenhouse', 'lever', 'instahyre', 'career_page', 'ats_crawler',
            'glassdoor', 'wellfound', 'workday', 'ashby',
        }
        valid_categories = {
            'marketing', 'finance', 'strategy', 'consulting', 'operations',
            'product-management', 'analytics', 'human-resources', 'supply-chain',
            'general-management', 'business-analytics',
            # NOTE: 'business-development' removed — all BD roles are auto-excluded
        }
        location_keywords = {
            'mumbai', 'delhi', 'bangalore', 'bengaluru', 'hyderabad',
            'chennai', 'pune', 'kolkata', 'gurugram', 'gurgaon', 'noida',
            'remote', 'wfh', 'ahmedabad', 'jaipur', 'lucknow', 'kochi',
        }

        if context.args:
            for arg in context.args:
                arg_lower = arg.lower().strip()

                if arg_lower.isdigit():
                    page = max(1, int(arg_lower))
                    continue
                if arg_lower in valid_sorts:
                    sort_by = arg_lower
                    continue
                if arg_lower in valid_sources:
                    source_filter = arg_lower
                    continue
                if arg_lower in valid_categories:
                    category_filter = arg_lower
                    continue
                if arg_lower in location_keywords:
                    location_filter = arg_lower
                    continue

                # Duration: "5mo" or "5m"
                dur_match = _re.match(r'^(\d+)\s*m(?:o(?:nths?)?)?$', arg_lower)
                if dur_match:
                    max_duration = max(1, min(12, int(dur_match.group(1))))
                    continue

                # Min stipend: "10k" or "10000"
                stip_match = _re.match(r'^(\d+)\s*k$', arg_lower)
                if stip_match:
                    min_stipend = int(stip_match.group(1)) * 1000
                    continue

                # Might be a category or location we don't know
                if len(arg_lower) > 2:
                    category_filter = arg_lower

        offset = (page - 1) * per_page

        await update.message.reply_text(
            f"🔍 Finding management internships...\n"
            f"⏱ ≤{max_duration}mo │ 🔀 by {sort_by} │ 🚫 No sales/BD"
        )

        listings, total = self.db.get_management_internships(
            limit=per_page,
            offset=offset,
            max_duration_months=max_duration,
            sort_by=sort_by,
            category=category_filter,
            source=source_filter,
            min_stipend=min_stipend,
            location=location_filter,
        )

        if not listings:
            msg_parts = ["📭 No management internships found with current filters."]
            if max_duration < 6:
                msg_parts.append(f"Try: /jobs 6mo (expand to 6 months)")
            if min_stipend > 0:
                msg_parts.append(f"Try: /jobs (remove stipend filter)")
            msg_parts.append("Or: /run pipeline to scrape new listings")
            await update.message.reply_text('\n'.join(msg_parts))
            return

        total_pages = max(1, (total + per_page - 1) // per_page)

        filters = {
            'category': category_filter,
            'source': source_filter,
            'location': location_filter,
            'min_stipend': min_stipend if min_stipend > 0 else None,
        }
        header = self.formatter.format_jobs_header(
            total, page, total_pages, sort_by, max_duration,
            {k: v for k, v in filters.items() if v}
        )

        lines = [header]
        start_num = offset + 1
        for i, l in enumerate(listings, start_num):
            lines.append(self.formatter._format_listing_line(i, l))

        lines.append(f"{'━' * 28}")

        # Navigation
        nav = []
        if page > 1:
            nav.append(f"⬅️ /jobs {sort_by} {page - 1}")
        if page < total_pages:
            nav.append(f"➡️ /jobs {sort_by} {page + 1}")
        if nav:
            lines.append(f"📄 {' │ '.join(nav)}")

        lines.append("")
        lines.append(
            "💡 <b>Shortcuts:</b>\n"
            "  /jobs ppo — Sort by PPO score\n"
            "  /jobs marketing — Filter category\n"
            "  /jobs mumbai — Filter location\n"
            "  /jobs 10k — Min ₹10,000 stipend\n"
            "  /jobs 6mo — Expand to 6 months"
        )

        await self._send_long_message(update, '\n'.join(lines))

    @command_error_boundary
    async def _cmd_ocean(self, update, context):
        """Blue Ocean listings with professional formatting."""
        listings = self.db.get_blue_ocean_listings(limit=15)
        if not listings:
            await update.message.reply_text(
                "🌊 No Blue Ocean listings found yet.\n"
                "Criteria: High prestige company + Low applicants (<35)\n"
                "Run /run pipeline to discover new opportunities."
            )
            return

        lines = [
            "🌊 <b>Blue Ocean Listings</b>",
            "<i>High prestige companies + Low competition</i>",
            f"<i>{datetime.now(IST).strftime('%d %b %Y, %I:%M %p IST')}</i>",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", "",
        ]

        for i, l in enumerate(listings, 1):
            lines.append(self.formatter._format_listing_line(i, l))

        lines.append("💡 These have the best effort-to-reward ratio. Apply NOW!")
        lines.append("/export for Excel with all details")
        await self._send_long_message(update, '\n'.join(lines))

    @command_error_boundary
    async def _cmd_internshala(self, update, context):
        """Live Internshala search."""
        if not context.args:
            await update.message.reply_text(
                "Usage: /internshala <query>\n"
                "Example: /internshala digital marketing\n\n"
                f"Categories: {', '.join(MBA_CATEGORIES[:5])}..."
            )
            return

        query = ' '.join(context.args)
        await update.message.reply_text(f"🔍 Searching Internshala for: {query}...")

        try:
            from agents.a03_primary_scraper import get_primary_scraper
            scraper = get_primary_scraper()
            listings = scraper.search_on_demand(query)

            if not listings:
                await update.message.reply_text(f"No results found for '{query}'")
                return

            lines = [f"🔍 <b>Internshala: '{query}'</b> — {len(listings)} results\n"]
            for i, l in enumerate(listings[:10], 1):
                lines.append(
                    f"{i}. <b>{l.title}</b> @ {l.company}\n"
                    f"   📍 {l.location} | 💰 {l.stipend}"
                )

            await self._send_long_message(update, '\n'.join(lines))
        except Exception as e:
            await update.message.reply_text(f"❌ Search failed: {e}")

    @command_error_boundary
    async def _cmd_dark(self, update, context):
        """Dark channel finds."""
        try:
            from agents.a02_dark_channel import get_dark_channel_listener
            listener = get_dark_channel_listener()
            listings = listener.get_recent_finds(days=3, limit=15)
            msg = listener.format_dark_report(listings)
            await self._send_long_message(update, msg)
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")

    @command_error_boundary
    async def _cmd_signals(self, update, context):
        """Active intent signals."""
        try:
            from agents.a01_intent_scanner import get_intent_scanner
            scanner = get_intent_scanner()
            msg = scanner.get_signal_report(days=7)
            await self._send_long_message(update, msg)
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")

    @command_error_boundary
    async def _cmd_package(self, update, context):
        """Generate full application package."""
        lid = self._parse_listing_id(context.args)
        if lid is None:
            await update.message.reply_text("Usage: /package <listing_id>\nFind IDs with /top or /ocean")
            return

        listing = self.db.get_clean_listing_by_id(lid)
        if not listing:
            await update.message.reply_text(f"❌ Listing #{lid} not found")
            return

        await update.message.reply_text(f"📦 Generating application package for #{lid}... (30-60s)")

        profile = {
            'college': self.db.get_setting('college', 'a top MBA program'),
            'specialization': self.db.get_setting('specialization', 'Marketing'),
        }

        try:
            response = self.router.generate_package(listing, profile)
            if response.success:
                title = listing.get('title', '')
                company = listing.get('company', '')
                msg = (
                    f"📦 <b>Application Package</b>\n"
                    f"<b>{title}</b> @ {company}\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                    f"{response.content[:3800]}"
                )
                await self._send_long_message(update, msg)
            else:
                await update.message.reply_text(f"❌ Generation failed: {response.error}")
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")

    @command_error_boundary
    async def _cmd_ats(self, update, context):
        """ATS keyword simulation."""
        lid = self._parse_listing_id(context.args)
        if lid is None:
            await update.message.reply_text("Usage: /ats <listing_id>")
            return

        await update.message.reply_text(f"🔬 Running ATS simulation for #{lid}...")

        try:
            from agents.a10_ats_simulator import get_ats_simulator
            sim = get_ats_simulator()
            result = sim.simulate(lid)

            if hasattr(result, 'to_telegram_msg'):
                await self._send_long_message(update, result.to_telegram_msg())
            elif isinstance(result, dict) and 'error' in result:
                await update.message.reply_text(f"❌ {result['error']}")
            else:
                msg = f"🔬 ATS Match: {result.match_percentage:.0f}%\n"
                msg += f"Missing: {', '.join(result.missing_keywords[:10])}"
                await update.message.reply_text(msg)
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")

    @command_error_boundary
    async def _cmd_cover(self, update, context):
        """Generate cover letter using A-13 engine."""
        lid = self._parse_listing_id(context.args)
        if lid is None:
            await update.message.reply_text("Usage: /cover <listing_id>")
            return

        listing = self.db.get_clean_listing_by_id(lid)
        if not listing:
            await update.message.reply_text(f"❌ Listing #{lid} not found")
            return

        await update.message.reply_text(f"✍️ Generating cover letter for #{lid}...")

        try:
            from agents.a13_auto_apply import get_auto_apply_orchestrator
            orchestrator = get_auto_apply_orchestrator()
            cover_letter = orchestrator.generate_cover_letter_only(lid)

            if cover_letter:
                company = listing.get('company', '')
                title = listing.get('title', '')
                msg = (
                    f"✍️ <b>Cover Letter</b>\n"
                    f"<b>{title}</b> @ {company}\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                    f"{cover_letter[:3800]}\n\n"
                    f"💡 /queue {lid} to add to auto-apply queue"
                )
                await self._send_long_message(update, msg)
            else:
                await update.message.reply_text(
                    f"❌ Failed to generate cover letter for #{lid}. "
                    f"Check AI router quota with /quota."
                )
        except Exception as e:
            logger.error(f"[{AGENT_ID}] Cover letter generation error for #{lid}: {e}")
            await update.message.reply_text(f"❌ Error generating cover letter: {e}")

    @command_error_boundary
    async def _cmd_network(self, update, context):
        """Alumni/network mapping."""
        if not context.args:
            await update.message.reply_text("Usage: /network <company name>")
            return

        company = ' '.join(context.args)
        await update.message.reply_text(f"🔗 Mapping network for {company}...")

        try:
            from agents.a09_network_mapper import get_network_mapper
            mapper = get_network_mapper()
            college = self.db.get_setting('college', '')
            spec = self.db.get_setting('specialization', '')
            result = mapper.map_network(company, college, spec)

            if hasattr(result, 'to_telegram_msg'):
                await self._send_long_message(update, result.to_telegram_msg())
            else:
                await update.message.reply_text(
                    f"🔗 Network for {company}: {result.get('alumni_found', 0)} alumni found"
                )
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")

    @command_error_boundary
    async def _cmd_apply(self, update, context):
        """Mark listing as applied."""
        lid = self._parse_listing_id(context.args)
        if lid is None:
            await update.message.reply_text("Usage: /apply <listing_id>")
            return

        listing = self.db.get_clean_listing_by_id(lid)
        if not listing:
            await update.message.reply_text(f"❌ Listing #{lid} not found")
            return

        outcome = Outcome(
            listing_id=lid,
            company_id=listing.get('company_id'),
            status='applied',
            ppo_score_at_apply=listing.get('ppo_score', 0),
        )
        self.db.insert_outcome(outcome)
        self.db.update_clean_listing_scores(lid, status='applied')

        title = listing.get('title', '')
        company = listing.get('company', '')
        await update.message.reply_text(
            f"✅ <b>Marked as Applied!</b>\n"
            f"{title} @ {company}\n\n"
            f"Track: /outcome {lid} interview (when called)\n"
            f"Or: /outcome {lid} rejected | /outcome {lid} offer",
            parse_mode='HTML'
        )

    @command_error_boundary
    async def _cmd_outcome(self, update, context):
        """Log application outcome."""
        if len(context.args) < 2:
            await update.message.reply_text(
                "Usage: /outcome <listing_id> <result>\n\n"
                f"Valid results: {', '.join(VALID_OUTCOMES)}\n\n"
                "Example: /outcome 42 interview"
            )
            return

        lid = self._parse_listing_id([context.args[0]])
        if lid is None:
            await update.message.reply_text("❌ Invalid listing ID")
            return

        status = context.args[1].lower()
        if status not in VALID_OUTCOMES:
            await update.message.reply_text(
                f"❌ Invalid status '{status}'\n"
                f"Valid: {', '.join(VALID_OUTCOMES)}"
            )
            return

        notes = ' '.join(context.args[2:]) if len(context.args) > 2 else ''

        outcome = Outcome(listing_id=lid, status=status, notes=notes)
        self.db.insert_outcome(outcome)

        emoji = {
            'applied': '📝', 'shortlisted': '📋', 'interview': '🎤',
            'rejected': '❌', 'offer': '🎉', 'ppo': '🏆', 'withdrawn': '🔙',
        }.get(status, '📝')

        await update.message.reply_text(
            f"{emoji} Outcome logged: <b>{status.upper()}</b> for #{lid}",
            parse_mode='HTML'
        )

    @command_error_boundary
    async def _cmd_cirs(self, update, context):
        """Company Intern Readiness Score."""
        if not context.args:
            await update.message.reply_text("Usage: /cirs <company name>")
            return

        company_name = ' '.join(context.args)
        company = self.db.fuzzy_match_company(company_name)

        if not company:
            await update.message.reply_text(f"❌ Company '{company_name}' not found")
            return

        name = company.get('name', '')
        tier = company.get('tier', 5)
        sector = company.get('sector', 'Unknown')
        cirs = company.get('cirs', 40)
        ats = company.get('ats_platform', '') or 'Unknown'
        city = company.get('hq_city', '') or 'Unknown'

        tier_name = {1: 'Elite', 2: 'Strong MNC', 3: 'Unicorn', 4: 'Startup', 5: 'Niche'}.get(tier, '?')

        msg = (
            f"🏢 <b>CIRS: {name}</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Tier: {tier} ({tier_name})\n"
            f"Sector: {sector}\n"
            f"HQ: {city}\n"
            f"ATS: {ats}\n\n"
            f"📊 <b>CIRS Score: {cirs:.0f}/100</b>\n\n"
            f"Components:\n"
            f"  • Intent Signal Strength: —\n"
            f"  • Historical PPO Rate: —\n"
            f"  • Glassdoor Rating: —\n"
            f"  • Funding Recency: —\n"
            f"  • Posting Frequency: —\n"
            f"  • Career Page Health: —\n"
            f"\n💡 CIRS updates after each A-01 signal scan"
        )
        await update.message.reply_text(msg, parse_mode='HTML')

    @command_error_boundary
    async def _cmd_research(self, update, context):
        """Full company research brief."""
        if not context.args:
            await update.message.reply_text("Usage: /research <company name>")
            return

        company = ' '.join(context.args)
        await update.message.reply_text(f"🔍 Researching {company}... (30-60s)")

        try:
            response = self.router.research_company(company)
            if response.success:
                msg = (
                    f"🔍 <b>Research: {company}</b>\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                    f"{response.content[:3800]}"
                )
                await self._send_long_message(update, msg)
            else:
                await update.message.reply_text(f"❌ Failed: {response.error}")
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")

    @command_error_boundary
    async def _cmd_stats(self, update, context):
        """Weekly funnel stats."""
        try:
            stats = self.db.get_weekly_stats()
            msg = self.formatter.stats_report(stats)
            await self._send_long_message(update, msg)
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")

    @command_error_boundary
    async def _cmd_health(self, update, context):
        """Agent health dashboard."""
        heartbeats = self.db.get_all_heartbeats()
        msg = self.formatter.health_report(heartbeats)
        await self._send_long_message(update, msg)

    @command_error_boundary
    async def _cmd_quota(self, update, context):
        """API quota usage including SerpAPI, Groq, Cerebras, DDG."""
        try:
            report = self.router.get_quota_report()
            await update.message.reply_text(report, parse_mode='HTML')
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")

    @command_error_boundary
    async def _cmd_export(self, update, context):
        """Export top listings as Excel file sent via Telegram."""
        n = 50
        if context.args:
            try:
                n = int(context.args[0])
            except ValueError:
                pass

        listings = self.db.get_top_listings(n=min(n, 200))
        if not listings:
            await update.message.reply_text("📤 No listings to export")
            return

        await update.message.reply_text(f"📊 Generating Excel report with {len(listings)} listings...")

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import tempfile

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "MBA Internship Listings"

            # Styles
            header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            ppo_fill = PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid')
            bo_fill = PatternFill(start_color='E6F0FF', end_color='E6F0FF', fill_type='solid')

            # Headers
            headers = [
                'Rank', 'Title', 'Company', 'Location', 'PPO Score',
                'Stipend (INR/mo)', 'Duration (mo)', 'Applicants',
                'PPO?', 'WFH?', 'Blue Ocean?', 'Source', 'Category',
                'Company Tier', 'Sector', 'URL'
            ]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # Data rows
            tier_labels = {1: 'T1 Elite', 2: 'T2 MNC', 3: 'T3 Unicorn', 4: 'T4 Startup', 5: 'T5 Niche'}
            for row_idx, l in enumerate(listings, 2):
                tier = l.get('tier')
                is_bo = l.get('is_blue_ocean', False)
                ppo_score = l.get('ppo_score', 0)

                row_data = [
                    row_idx - 1,
                    l.get('title', ''),
                    l.get('company', ''),
                    l.get('location', ''),
                    round(ppo_score, 1),
                    l.get('stipend_monthly', 0) or 0,
                    l.get('duration_months', 0) or 0,
                    l.get('applicants', 0) or 0,
                    'Yes' if l.get('is_ppo') else 'No',
                    'Yes' if l.get('is_wfh') else 'No',
                    'Yes' if is_bo else 'No',
                    l.get('source', ''),
                    l.get('category', ''),
                    tier_labels.get(tier, 'Unknown') if tier else 'Unknown',
                    l.get('sector', ''),
                    l.get('url', ''),
                ]

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if col_idx == 5 and ppo_score >= 70:  # High PPO highlight
                        cell.fill = ppo_fill
                    if is_bo:
                        cell.fill = bo_fill

            # Auto-fit column widths
            col_widths = [6, 40, 25, 20, 10, 14, 12, 12, 6, 6, 10, 14, 18, 14, 14, 50]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Auto-filter
            ws.auto_filter.ref = ws.dimensions

            # Save to temp file
            date_str = datetime.now(IST).strftime('%Y-%m-%d_%H%M')
            filename = f"MBA_Listings_Top{len(listings)}_{date_str}.xlsx"

            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                wb.save(tmp.name)
                tmp_path = tmp.name

            # Send file via Telegram
            with open(tmp_path, 'rb') as f:
                await update.message.reply_document(
                    document=f,
                    filename=filename,
                    caption=(
                        f"📊 <b>MBA Internship Report</b>\n"
                        f"Generated: {datetime.now(IST).strftime('%d %b %Y %I:%M %p IST')}\n"
                        f"Listings: {len(listings)} | Sorted by PPO Score"
                    ),
                    parse_mode='HTML'
                )

            # Cleanup
            import os as _os
            _os.unlink(tmp_path)

        except ImportError:
            # openpyxl not available, fall back to text export
            lines = ["RANK | TITLE | COMPANY | PPO | STIPEND | LOCATION | SOURCE | URL"]
            lines.append("-" * 100)
            for i, l in enumerate(listings, 1):
                lines.append(
                    f"{i} | {l.get('title', '')[:30]} | {l.get('company', '')[:20]} | "
                    f"{l.get('ppo_score', 0):.1f} | {l.get('stipend_monthly', 0) or 0:,.0f} | "
                    f"{l.get('location', '')[:15]} | {l.get('source', '')} | {l.get('url', '')[:50]}"
                )
            text = '\n'.join(lines)
            for i in range(0, len(text), TG_MAX_LEN):
                await update.message.reply_text(text[i:i + TG_MAX_LEN])

        except Exception as e:
            logger.error(f"[{AGENT_ID}] Export error: {e}")
            await update.message.reply_text(f"❌ Export failed: {e}")

    @command_error_boundary
    async def _cmd_settings(self, update, context):
        """User preferences."""
        if not context.args:
            college = self.db.get_setting('college', 'Not set')
            spec = self.db.get_setting('specialization', 'Not set')
            resume = 'Set ✅' if self.db.get_setting('user_resume', '') else 'Not set ❌'

            msg = (
                f"⚙️ <b>Settings</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"🎓 College: {college}\n"
                f"📚 Specialization: {spec}\n"
                f"📄 Resume: {resume}\n\n"
                f"<b>Set with:</b>\n"
                f"/settings college Your College Name\n"
                f"/settings spec Marketing\n"
                f"/settings resume [paste your resume text]"
            )
            await update.message.reply_text(msg, parse_mode='HTML')
            return

        key = context.args[0].lower()
        value = ' '.join(context.args[1:])

        if not value:
            await update.message.reply_text(f"Usage: /settings {key} <value>")
            return

        key_map = {
            'college': 'college',
            'spec': 'specialization',
            'specialization': 'specialization',
            'resume': 'user_resume',
        }

        db_key = key_map.get(key)
        if not db_key:
            await update.message.reply_text(
                f"❌ Unknown setting '{key}'\n"
                f"Available: college, spec, resume"
            )
            return

        self.db.set_setting(db_key, value)
        await update.message.reply_text(f"✅ <b>{key}</b> updated!", parse_mode='HTML')

    @command_error_boundary
    async def _cmd_refresh(self, update, context):
        """Force re-scrape. Delegates to /run pipeline for non-blocking execution."""
        await update.message.reply_text(
            "🔄 Starting full pipeline refresh in background...\n"
            "Use /status to monitor progress."
        )
        # Delegate to the /run pipeline system which handles
        # async/sync properly and streams progress
        context.args = ['pipeline']
        await self._cmd_run(update, context)

    # ================================================================
    # /run COMMAND — MANUAL AGENT TRIGGER WITH STREAMING PROGRESS
    # ================================================================

    # Map of runnable agent names to their runner info
    _RUN_AGENTS = {
        'pipeline': {
            'desc': 'Full pipeline: scrape → dedup → ghost → enrich → PPO → brief',
            'steps': ['scrape', 'dedup', 'ghost', 'enrich', 'ppo', 'brief'],
        },
        'scrape': {
            'desc': 'A-03: Internshala full scrape (10 categories)',
            'agent': 'A-03',
        },
        'afternoon': {
            'desc': 'A-03: Naukri + IIMjobs afternoon scrape',
            'agent': 'A-03',
        },
        'dedup': {
            'desc': 'A-06: Deduplication engine',
            'agent': 'A-06',
        },
        'ghost': {
            'desc': 'A-05: Ghost job detection & scoring',
            'agent': 'A-05',
        },
        'enrich': {
            'desc': 'A-07: Intelligence enrichment + CIRS',
            'agent': 'A-07',
        },
        'ppo': {
            'desc': 'A-08: PPO scoring & ranking',
            'agent': 'A-08',
        },
        'intent': {
            'desc': 'A-01: Hiring intent signal scan',
            'agent': 'A-01',
        },
        'ats_crawl': {
            'desc': 'A-04: Company ATS career page crawl',
            'agent': 'A-04',
        },
        'dark': {
            'desc': 'A-02: Dark channel batch check',
            'agent': 'A-02',
        },
        'brief': {
            'desc': 'A-12: Generate & send morning brief',
            'agent': 'A-12',
        },
    }

    @command_error_boundary
    async def _cmd_run(self, update, context):
        """
        Run an agent or the full pipeline manually.
        ADMIN ONLY — only admins can trigger pipeline/scrape operations.
        Runs in the BACKGROUND so Telegram stays responsive.
        Streams real-time progress updates to the chat.
        """
        # Admin-only check for pipeline control
        telegram_id = update.effective_user.id if update.effective_user else 0
        if not self.security.is_admin(telegram_id):
            await update.message.reply_text(
                "🔒 Pipeline control is admin-only.\n"
                "Contact the admin to run agents."
            )
            return

        if not context.args:
            lines = [
                "🚀 <b>Manual Agent Runner</b>",
                "━━━━━━━━━━━━━━━━━━━━━━━━━━━",
                "",
                "Usage: <code>/run agent_name</code>",
                "",
                "⚡ <b>Full Pipeline:</b>",
                "  <code>/run pipeline</code> — Runs ALL steps in order",
                "",
                "🔧 <b>Individual Agents:</b>",
            ]
            for name, info in self._RUN_AGENTS.items():
                if name == 'pipeline':
                    continue
                lines.append(f"  <code>/run {name}</code> — {info['desc']}")

            lines.append("")
            lines.append("📊 <b>Background Task Control:</b>")
            lines.append("  <code>/status</code> — See running tasks")
            lines.append("  <code>/cancel task_name</code> — Cancel a task")
            lines.append("")
            lines.append("💡 Agents run in background. You'll get live progress updates.")
            lines.append("💡 Scheduled cycles are NOT affected — they run at their normal times.")
            await update.message.reply_text('\n'.join(lines), parse_mode='HTML')
            return

        agent_name = context.args[0].lower()

        if agent_name not in self._RUN_AGENTS:
            await update.message.reply_text(
                f"❌ Unknown agent '{agent_name}'\n"
                f"Available: {', '.join(self._RUN_AGENTS.keys())}\n"
                f"Use /run for full list."
            )
            return

        # Check if this agent is already running
        async with self._task_lock:
            if agent_name in self._running_tasks:
                task_info = self._running_tasks[agent_name]
                elapsed = time.time() - task_info['start_time']
                await update.message.reply_text(
                    f"⚠️ <b>{agent_name}</b> is already running "
                    f"({elapsed:.0f}s elapsed)\n"
                    f"Use /cancel {agent_name} to stop it first.",
                    parse_mode='HTML'
                )
                return

        info = self._RUN_AGENTS[agent_name]
        chat_id = update.effective_chat.id

        # Pipeline mode: run multiple agents in sequence
        if 'steps' in info:
            task = asyncio.create_task(
                self._run_pipeline_bg(chat_id, agent_name, info['steps'])
            )
        else:
            task = asyncio.create_task(
                self._run_single_agent_bg(chat_id, agent_name, info)
            )

        async with self._task_lock:
            self._running_tasks[agent_name] = {
                'task': task,
                'start_time': time.time(),
                'desc': info['desc'],
                'chat_id': chat_id,
                'status': 'running',
            }

        await update.message.reply_text(
            f"🏃 <b>{info['desc']}</b> — started in background\n"
            f"You'll get live progress updates here.\n"
            f"Use /status to check | <code>/cancel {agent_name}</code> to stop",
            parse_mode='HTML'
        )

    async def _stream_msg(self, chat_id: int, text: str):
        """Send a progress message to a specific chat. Falls back to plain text."""
        try:
            if self._app and self._app.bot:
                try:
                    await self._app.bot.send_message(
                        chat_id=chat_id,
                        text=text,
                        parse_mode='HTML'
                    )
                except Exception:
                    # Fallback: strip HTML and send as plain text
                    import re
                    plain = re.sub(r'<[^>]+>', '', text)
                    try:
                        await self._app.bot.send_message(
                            chat_id=chat_id,
                            text=plain,
                        )
                    except Exception as e2:
                        logger.debug(f"[{AGENT_ID}] Stream msg fallback failed: {e2}")
        except Exception as e:
            logger.debug(f"[{AGENT_ID}] Stream msg failed: {e}")

    async def _run_single_agent_bg(self, chat_id: int, agent_name: str, info: Dict):
        """Background task: Run a single agent with streaming updates."""
        start_time = time.time()
        try:
            result = await self._execute_agent(agent_name)
            duration = time.time() - start_time

            # Format result
            if isinstance(result, dict):
                result_str = '\n'.join(
                    f"  {k}: {v}" for k, v in result.items()
                    if k not in ('raw', 'listings', 'html')
                )
            elif result is not None:
                result_str = str(result)[:500]
            else:
                result_str = "  (completed, no detailed output)"

            await self._stream_msg(
                chat_id,
                f"✅ <b>{info['desc']}</b> — DONE\n"
                f"⏱ Duration: {duration:.1f}s\n\n"
                f"📊 Results:\n{result_str}"
            )

        except asyncio.CancelledError:
            duration = time.time() - start_time
            await self._stream_msg(
                chat_id,
                f"🛑 <b>{info['desc']}</b> — CANCELLED\n"
                f"⏱ Ran for: {duration:.1f}s"
            )
        except Exception as e:
            duration = time.time() - start_time
            logger.error(f"[{AGENT_ID}] /run {agent_name} failed: {e}")
            tb_str = traceback.format_exc()[-300:]
            await self._stream_msg(
                chat_id,
                f"❌ <b>{info['desc']}</b> — FAILED\n"
                f"⏱ Duration: {duration:.1f}s\n"
                f"Error: {str(e)[:200]}\n\n"
                f"<code>{tb_str}</code>"
            )
        finally:
            async with self._task_lock:
                self._running_tasks.pop(agent_name, None)

    async def _run_pipeline_bg(self, chat_id: int, task_name: str, steps: list):
        """Background task: Run full pipeline with per-step streaming."""
        total_start = time.time()
        await self._stream_msg(
            chat_id,
            "🚀 <b>FULL PIPELINE STARTING</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"Steps: {' → '.join(steps)}\n\n"
            "You'll get a progress update after each step."
        )

        results = {}
        failed = []

        for i, step in enumerate(steps, 1):
            info = self._RUN_AGENTS.get(step, {})
            desc = info.get('desc', step)
            step_start = time.time()

            await self._stream_msg(
                chat_id,
                f"⏳ [{i}/{len(steps)}] <b>{desc}</b>..."
            )

            try:
                result = await self._execute_agent(step)
                duration = time.time() - step_start

                # Extract summary
                if isinstance(result, dict):
                    summary = ', '.join(
                        f"{k}={v}" for k, v in result.items()
                        if k in ('total', 'new', 'processed', 'enriched',
                                 'scored', 'signals', 'duplicates', 'blue_ocean')
                    ) or 'OK'
                else:
                    summary = 'OK'

                results[step] = f"✅ {duration:.0f}s — {summary}"

                # Stream per-step completion
                await self._stream_msg(
                    chat_id,
                    f"  ✅ [{i}/{len(steps)}] {step}: {duration:.0f}s — {summary}"
                )

            except asyncio.CancelledError:
                results[step] = f"🛑 CANCELLED"
                failed.append(step)
                break
            except Exception as e:
                duration = time.time() - step_start
                results[step] = f"❌ {duration:.0f}s — {str(e)[:80]}"
                failed.append(step)
                logger.error(f"[{AGENT_ID}] Pipeline step '{step}' failed: {e}")
                # Continue with next step — partial processing is better than none

        # Final summary
        total_duration = time.time() - total_start
        lines = [
            f"🏁 <b>PIPELINE COMPLETE</b> — {total_duration:.0f}s",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "",
        ]
        for step, result_str in results.items():
            lines.append(f"  {step}: {result_str}")

        if failed:
            lines.append(f"\n⚠️ {len(failed)} step(s) failed: {', '.join(failed)}")
        else:
            lines.append(f"\n🎉 All {len(steps)} steps completed!")

        lines.append(f"\n💡 Use /top or /morning to see results.")
        await self._stream_msg(chat_id, '\n'.join(lines))

        async with self._task_lock:
            self._running_tasks.pop(task_name, None)

    async def _execute_agent(self, agent_name: str):
        """
        Execute a specific agent and return its result.
        Handles both sync and async agent methods seamlessly.
        """
        loop = asyncio.get_running_loop()

        if agent_name == 'scrape':
            from agents.a03_primary_scraper import get_primary_scraper
            # run_morning_scrape is SYNC — run in executor to not block
            return await loop.run_in_executor(
                None, get_primary_scraper().run_morning_scrape
            )

        elif agent_name == 'afternoon':
            from agents.a03_primary_scraper import get_primary_scraper
            return await loop.run_in_executor(
                None, get_primary_scraper().run_afternoon_scrape
            )

        elif agent_name == 'dedup':
            from agents.a06_dedup_engine import get_dedup_engine
            return await loop.run_in_executor(
                None, get_dedup_engine().run_dedup
            )

        elif agent_name == 'ghost':
            from agents.a05_ghost_detector import get_ghost_detector
            return await loop.run_in_executor(
                None, get_ghost_detector().score_batch
            )

        elif agent_name == 'enrich':
            from agents.a07_intelligence_enricher import get_intelligence_enricher
            return await loop.run_in_executor(
                None, get_intelligence_enricher().run_enrichment
            )

        elif agent_name == 'ppo':
            from agents.a08_ppo_optimizer import get_ppo_optimizer
            return await loop.run_in_executor(
                None, get_ppo_optimizer().run_optimization
            )

        elif agent_name == 'intent':
            from agents.a01_intent_scanner import get_intent_scanner
            return await loop.run_in_executor(
                None, get_intent_scanner().run_scan
            )

        elif agent_name == 'ats_crawl':
            from agents.a04_ats_crawler import get_ats_crawler
            return await loop.run_in_executor(
                None, get_ats_crawler().run_crawl
            )

        elif agent_name == 'dark':
            from agents.a02_dark_channel import get_dark_channel_listener
            return await loop.run_in_executor(
                None, get_dark_channel_listener().run_batch_check
            )

        elif agent_name == 'brief':
            await self.send_morning_brief()
            return {'status': 'sent'}

        else:
            raise ValueError(f"Unknown agent: {agent_name}")

    # ================================================================
    # /status COMMAND — SHOW RUNNING BACKGROUND TASKS
    # ================================================================

    @command_error_boundary
    async def _cmd_status(self, update, context):
        """Show currently running background tasks."""
        async with self._task_lock:
            if not self._running_tasks:
                await update.message.reply_text(
                    "💤 No background tasks running.\n"
                    "Use /run to start an agent."
                )
                return

            lines = [
                "🏃 <b>Running Background Tasks</b>",
                "━━━━━━━━━━━━━━━━━━━━━━━━━━━",
                "",
            ]
            for name, info in self._running_tasks.items():
                elapsed = time.time() - info['start_time']
                desc = info.get('desc', name)
                lines.append(
                    f"  🔄 <b>{name}</b>: {desc}\n"
                    f"     Running for {elapsed:.0f}s"
                )
            lines.append(f"\n💡 Use <code>/cancel task_name</code> to stop a task.")

        await update.message.reply_text('\n'.join(lines), parse_mode='HTML')

    # ================================================================
    # /cancel COMMAND — CANCEL A RUNNING TASK
    # ================================================================

    @command_error_boundary
    async def _cmd_cancel(self, update, context):
        """Cancel a running background task."""
        if not context.args:
            async with self._task_lock:
                if not self._running_tasks:
                    await update.message.reply_text("💤 No tasks running.")
                    return
                names = ', '.join(self._running_tasks.keys())
            await update.message.reply_text(
                f"Usage: /cancel <task_name>\n"
                f"Running tasks: {names}"
            )
            return

        task_name = context.args[0].lower()

        async with self._task_lock:
            task_info = self._running_tasks.get(task_name)
            if not task_info:
                await update.message.reply_text(
                    f"❌ No task '{task_name}' is running.\n"
                    f"Use /status to see running tasks."
                )
                return

            task = task_info.get('task')
            if task and not task.done():
                task.cancel()
                await update.message.reply_text(
                    f"🛑 Cancelling <b>{task_name}</b>...",
                    parse_mode='HTML'
                )
            else:
                self._running_tasks.pop(task_name, None)
                await update.message.reply_text(
                    f"Task '{task_name}' already finished."
                )

    # ================================================================
    # /queue COMMAND — ADD TO AUTO-APPLY QUEUE
    # ================================================================

    @command_error_boundary
    async def _cmd_queue(self, update, context):
        """Queue a listing for auto-apply or queue top N."""
        try:
            from agents.a13_auto_apply import get_auto_apply_orchestrator
            orchestrator = get_auto_apply_orchestrator()

            if not context.args:
                # Show queue status
                status_msg = orchestrator.get_queue_status()
                await self._send_long_message(update, status_msg)
                return

            arg = context.args[0].lower()

            if arg == 'top':
                # Queue top N by PPO score
                n = 10
                if len(context.args) > 1:
                    try:
                        n = int(context.args[1])
                    except ValueError:
                        pass
                queued = orchestrator.queue_manager.queue_top_listings(n=n)
                await update.message.reply_text(
                    f"📬 Queued {queued} top listings for auto-apply.\n"
                    f"Use /autoapply to start processing."
                )
            else:
                # Queue specific listing
                try:
                    lid = int(arg)
                except ValueError:
                    await update.message.reply_text("Usage: /queue <listing_id> or /queue top [N]")
                    return

                result = orchestrator.queue_and_confirm(lid)
                if 'error' in result:
                    await update.message.reply_text(f"❌ {result['error']}")
                    return

                msg = (
                    f"📬 <b>Queued for Auto-Apply</b>\n"
                    f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                    f"<b>{result['title']}</b>\n"
                    f"🏢 {result['company']}\n"
                    f"📡 Platform: {result['platform']}\n\n"
                    f"📝 <b>Cover Letter Preview:</b>\n"
                    f"<i>{result['cover_letter_preview'][:300]}...</i>\n\n"
                    f"Use /autoapply to submit queued applications."
                )
                await self._send_long_message(update, msg)

        except Exception as e:
            logger.error(f"[{AGENT_ID}] Queue error: {e}")
            await update.message.reply_text(f"❌ Error: {e}")

    # ================================================================
    # /autoapply COMMAND — RUN AUTO-APPLY SESSION
    # ================================================================

    @command_error_boundary
    async def _cmd_autoapply(self, update, context):
        """Run auto-apply on queued applications."""
        try:
            from agents.a13_auto_apply import get_auto_apply_orchestrator
            orchestrator = get_auto_apply_orchestrator()

            max_apps = 5
            if context.args:
                try:
                    max_apps = min(int(context.args[0]), 15)
                except ValueError:
                    pass

            await update.message.reply_text(
                f"🚀 Starting auto-apply (max {max_apps} applications)...\n"
                f"This may take a few minutes."
            )

            # Run in background
            import asyncio
            loop = asyncio.get_event_loop()
            stats = await loop.run_in_executor(
                None, lambda: orchestrator.run_auto_apply(max_apps=max_apps)
            )

            await self._send_long_message(update, stats.to_telegram_msg())

        except Exception as e:
            logger.error(f"[{AGENT_ID}] Auto-apply error: {e}")
            await update.message.reply_text(f"❌ Auto-apply error: {e}")

    # ================================================================
    # /appstatus COMMAND — APPLICATION HISTORY
    # ================================================================

    @command_error_boundary
    async def _cmd_appstatus(self, update, context):
        """Show application history and stats."""
        try:
            history = self.db.get_application_history(limit=15)
            stats = self.db.get_application_stats()

            lines = [
                f"📝 <b>Application History</b>",
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
                f"",
                f"✅ Applied: {stats.get('applied', 0)}",
                f"❌ Failed: {stats.get('failed', 0)}",
                f"📬 Queued: {stats.get('queued', 0)}",
                f"📊 Today: {stats.get('applied_today', 0)}",
                f"",
            ]

            if history:
                lines.append("<b>Recent Applications:</b>")
                for i, app in enumerate(history[:10], 1):
                    status_emoji = '✅' if app.get('status') == 'applied' else '❌'
                    title = app.get('title', '')[:35]
                    company = app.get('company', '')[:20]
                    applied_at = app.get('applied_at', '')[:10]
                    lines.append(
                        f"{i}. {status_emoji} {title} @ {company}"
                    )
                    if applied_at:
                        lines.append(f"   📅 {applied_at}")

            lines.append(f"\n💡 /queue to add | /autoapply to run")
            await self._send_long_message(update, '\n'.join(lines))

        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")

    # ================================================================
    # /schedule COMMAND — SHOW SCHEDULE
    # ================================================================

    @command_error_boundary
    async def _cmd_schedule(self, update, context):
        """Show the full 24-hour schedule with next run times."""
        try:
            from core.scheduler import get_scheduler
            scheduler = get_scheduler()
            msg = scheduler.get_schedule_display()

            # Also show next runs from APScheduler
            jobs = scheduler.get_job_list()
            if jobs:
                msg += "\n\n📅 <b>Next Scheduled Runs:</b>\n"
                for job in jobs[:15]:
                    name = job.get('name', '')
                    next_run = job.get('next_run', 'N/A')
                    msg += f"  {name}\n    → {next_run}\n"

            await self._send_long_message(update, msg)
        except Exception as e:
            await update.message.reply_text(f"❌ Error: {e}")

    # ================================================================
    # /loadall — BROWSE ALL LISTINGS WITH PAGINATION
    # ================================================================

    @command_error_boundary
    async def _cmd_loadall(self, update, context):
        """Browse ALL clean listings with pagination and filters.
        Usage: /loadall [page] [category] [source]
        Examples:
            /loadall              — Page 1, all listings
            /loadall 2            — Page 2
            /loadall 1 marketing  — Page 1, marketing only
            /loadall 1 all linkedin — Page 1, LinkedIn only
        """
        page = 1
        category = None
        source = None
        per_page = 15

        if context.args:
            # Parse page number
            try:
                page = max(1, int(context.args[0]))
            except ValueError:
                pass

            # Parse category filter
            if len(context.args) > 1:
                cat_arg = context.args[1].lower()
                if cat_arg != 'all':
                    category = cat_arg

            # Parse source filter
            if len(context.args) > 2:
                source = context.args[2].lower()

        offset = (page - 1) * per_page

        listings = self.db.get_all_clean_listings(
            limit=per_page, offset=offset,
            category=category, source=source,
            sort_by='ppo_score', sort_order='DESC'
        )

        if not listings:
            msg = "📋 No listings found"
            if category:
                msg += f" for category '{category}'"
            if source:
                msg += f" from '{source}'"
            msg += f" on page {page}."
            msg += "\n\nTry: /loadall 1"
            await update.message.reply_text(msg)
            return

        # Count totals for header
        total = self.db.count_clean_listings_filtered(
            category=category, source=source
        )
        total_pages = max(1, (total + per_page - 1) // per_page)

        lines = [
            f"📋 <b>ALL LISTINGS</b> — Page {page}/{total_pages} ({total} total)",
        ]
        if category or source:
            filters = []
            if category:
                filters.append(f"📂 {category}")
            if source:
                filters.append(f"📡 {source}")
            lines.append(f"<i>Filters: {' | '.join(filters)}</i>")
        lines.extend(["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""])

        start_num = offset + 1
        for i, l in enumerate(listings, start_num):
            lines.append(self.formatter._format_listing_line(i, l))

        # Navigation hints
        nav = []
        if page > 1:
            nav.append(f"/loadall {page-1}" + (f" {category}" if category else "") + (f" {source}" if source else ""))
        if page < total_pages:
            nav.append(f"/loadall {page+1}" + (f" {category}" if category else "") + (f" {source}" if source else ""))
        if nav:
            lines.append(f"\n📄 Navigate: {' | '.join(nav)}")

        lines.append(f"\n💡 /filter to see available filters")
        await self._send_long_message(update, '\n'.join(lines))

    # ================================================================
    # /filter — SHOW AVAILABLE FILTERS AND STATS
    # ================================================================

    @command_error_boundary
    async def _cmd_filter(self, update, context):
        """Show available filter categories and sources with counts.
        Usage: /filter
        """
        # Get category counts
        cat_counts = self.db.get_category_counts()
        source_counts = self.db.get_source_counts()

        lines = [
            "🔍 <b>FILTER OPTIONS</b>",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "",
            "📂 <b>By Category:</b>",
        ]

        for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
            if count > 0:
                lines.append(f"  • <code>{cat}</code> — {count} listings")
                lines.append(f"    → /loadall 1 {cat}")

        lines.extend(["", "📡 <b>By Source:</b>"])
        for src, count in sorted(source_counts.items(), key=lambda x: -x[1]):
            if count > 0:
                emoji = self.formatter.SOURCE_EMOJI.get(src, '📡')
                lines.append(f"  {emoji} <code>{src}</code> — {count} listings")
                lines.append(f"    → /loadall 1 all {src}")

        lines.extend([
            "",
            "📊 <b>Combine filters:</b>",
            "  /loadall 1 marketing linkedin",
            "  /loadall 1 finance internshala",
            "",
            "🏆 <b>Quick views:</b>",
            "  /top 20 — Top 20 by PPO score",
            "  /ocean — Blue Ocean (low competition)",
            "  /loadall — Browse all",
        ])

        await self._send_long_message(update, '\n'.join(lines))

    # ================================================================
    # /sources — SOURCE HEALTH & STATISTICS
    # ================================================================

    @command_error_boundary
    async def _cmd_sources(self, update, context):
        """Show detailed source health and scraping statistics."""
        source_counts = self.db.get_source_counts()
        raw_counts = self.db.get_raw_source_counts()

        lines = [
            "📡 <b>SOURCE HEALTH DASHBOARD</b>",
            f"<i>{datetime.now(IST).strftime('%d %b %Y, %I:%M %p IST')}</i>",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "",
        ]

        total_clean = sum(source_counts.values())
        total_raw = sum(raw_counts.values())

        lines.append(f"📊 Total clean listings: <b>{total_clean}</b>")
        lines.append(f"📦 Total raw scraped: <b>{total_raw}</b>")
        lines.append(f"🔄 Conversion rate: <b>{total_clean/max(total_raw,1)*100:.1f}%</b>")
        lines.append("")

        all_sources = set(list(source_counts.keys()) + list(raw_counts.keys()))
        for src in sorted(all_sources):
            emoji = self.formatter.SOURCE_EMOJI.get(src, '📡')
            clean = source_counts.get(src, 0)
            raw = raw_counts.get(src, 0)
            status = "✅" if clean > 0 else ("⚠️" if raw > 0 else "❌")
            lines.append(
                f"{status} {emoji} <b>{src}</b>: "
                f"{clean} clean / {raw} raw"
            )

        lines.extend([
            "",
            "💡 <b>Commands:</b>",
            "  /refresh — Force re-scrape all sources",
            "  /run pipeline — Run full processing pipeline",
            "  /loadall 1 all <source> — Browse by source",
        ])

        await self._send_long_message(update, '\n'.join(lines))

    # ================================================================
    # /browse — QUICK CATEGORY BROWSER
    # ================================================================

    @command_error_boundary
    async def _cmd_browse(self, update, context):
        """Quick category browser — jump to any category.
        Usage: /browse [category]
        """
        if not context.args:
            # Show category menu
            cat_counts = self.db.get_category_counts()
            lines = [
                "📂 <b>CATEGORY BROWSER</b>",
                "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
                "",
            ]
            for cat, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
                if count > 0:
                    lines.append(f"  📁 /browse {cat} — {count} listings")

            lines.extend([
                "",
                "💡 Tap any category above to view listings",
            ])
            await self._send_long_message(update, '\n'.join(lines))
            return

        # Show listings for that category
        category = context.args[0].lower()
        listings = self.db.get_all_clean_listings(
            limit=15, offset=0,
            category=category, sort_by='ppo_score'
        )

        if not listings:
            await update.message.reply_text(
                f"📂 No listings in '{category}'.\n"
                f"Try /browse to see available categories."
            )
            return

        total = self.db.count_clean_listings_filtered(category=category)
        lines = [
            f"📂 <b>{category.upper()}</b> — {total} listings (showing top 15)",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", "",
        ]
        for i, l in enumerate(listings, 1):
            lines.append(self.formatter._format_listing_line(i, l))

        if total > 15:
            lines.append(f"\n📄 More: /loadall 2 {category}")
        lines.append(f"💡 /export for Excel | /browse for categories")
        await self._send_long_message(update, '\n'.join(lines))

    # ================================================================
    # /cfstatus — CLOUDFLARE CRAWL STATUS
    # ================================================================

    @command_error_boundary
    async def _cmd_cfstatus(self, update, context):
        """Show Cloudflare /crawl API configuration status."""
        try:
            from core.cloudflare_crawl import get_status
            status = get_status()
        except ImportError:
            await update.message.reply_text("❌ Cloudflare crawl module not found")
            return

        configured = status.get('configured', False)
        lines = [
            "☁️ <b>CLOUDFLARE /CRAWL STATUS</b>",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "",
            f"Status: {'✅ CONFIGURED' if configured else '❌ NOT CONFIGURED'}",
            f"Worker URL: {status.get('worker_url', 'NOT SET')}",
            f"Requests this hour: {status.get('requests_this_hour', 0)}/{status.get('max_per_hour', 10)}",
            f"Free tier: {status.get('free_tier_limit', '5,000/month')}",
        ]
        if not configured:
            lines.extend([
                "",
                "📋 <b>Setup Guide (FREE):</b>",
                "1. Create Cloudflare account",
                "2. Enable Browser Rendering (free tier)",
                "3. Deploy Worker (see core/cloudflare_crawl.py)",
                "4. Set env vars:",
                "   CF_CRAWL_WORKER_URL=https://your.workers.dev",
                "   CF_CRAWL_SECRET=your-secret",
            ])
        await self._send_long_message(update, '\n'.join(lines))

    # ================================================================
    # /reprocess — FORCE RE-PROCESS STUCK RAW LISTINGS
    # ================================================================

    @command_error_boundary
    async def _cmd_reprocess(self, update, context):
        """Force re-process stuck/pending raw listings through the pipeline.
        Usage: /reprocess [reset]
            /reprocess       — Show status of raw listings
            /reprocess reset — Reset all 'duplicate' raw to 'pending' for re-eval
        """
        unprocessed = self.db.count_unprocessed_raw_listings()

        if context.args and context.args[0].lower() == 'reset':
            # Reset all duplicate raw listings back to pending
            with self.db.get_cursor() as cur:
                cur.execute(
                    "UPDATE raw_listings SET dedup_status = 'pending' "
                    "WHERE dedup_status = 'duplicate'"
                )
                reset_count = cur.rowcount

            await update.message.reply_text(
                f"🔄 Reset {reset_count} duplicate raw listings to pending.\n"
                f"They will be re-evaluated in the next dedup cycle.\n"
                f"Run /run pipeline to process them now."
            )
            return

        # Show raw listing status breakdown
        with self.db.get_cursor() as cur:
            cur.execute(
                "SELECT dedup_status, COUNT(*) as cnt "
                "FROM raw_listings GROUP BY dedup_status"
            )
            status_counts = {row['dedup_status']: row['cnt'] for row in cur.fetchall()}

        total_raw = sum(status_counts.values())
        lines = [
            "📦 <b>RAW LISTINGS STATUS</b>",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "",
            f"Total raw: <b>{total_raw}</b>",
        ]
        for status, count in sorted(status_counts.items()):
            emoji = {'pending': '⏳', 'new': '✅', 'duplicate': '🔄', 'filtered': '🚫'}.get(status, '❓')
            lines.append(f"  {emoji} {status}: <b>{count}</b>")

        lines.extend([
            "",
            "💡 <b>Commands:</b>",
            "  /reprocess reset — Reset duplicates for re-evaluation",
            "  /run pipeline — Process pending listings",
        ])
        await self._send_long_message(update, '\n'.join(lines))

    # ================================================================
    # /miniapp COMMAND — OPEN MINI APP
    # ================================================================

    @command_error_boundary
    async def _cmd_miniapp(self, update, context):
        """Open the InternHub Pro Mini App with inline keyboard button.
        
        Usage: /miniapp
        
        Sends a professional message with:
        1. WebApp button to open the mini app directly in Telegram
        2. User's access code for login
        3. Direct URL as fallback
        """
        telegram_id = update.effective_user.id if update.effective_user else 0
        
        if not MINI_APP_URL:
            await update.message.reply_text(
                "⚠️ <b>Mini App Not Configured</b>\n\n"
                "The InternHub Pro mini app URL has not been set.\n"
                "Ask the admin to configure <code>MINI_APP_URL</code> in environment variables.\n\n"
                "💡 In the meantime, use these bot commands:\n"
                "/jobs — Browse filtered internships\n"
                "/top 20 — Top 20 by PPO score\n"
                "/morning — Full morning brief",
                parse_mode='HTML'
            )
            return
        
        # Get user's access code
        user_data = self.security.get_user(telegram_id)
        access_code = user_data.get('access_code', '') if user_data else ''
        username = user_data.get('username', '') if user_data else ''
        
        msg = (
            "📱 <b>InternHub Pro — Mini App</b>\n"
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            "🚀 <b>Features:</b>\n"
            "  • Smart job browser with filters\n"
            "  • AI-powered application packages\n"
            "  • Real-time analytics dashboard\n"
            "  • Batch apply functionality\n"
            "  • Company research & CIRS scores\n\n"
        )
        
        if access_code:
            msg += (
                "🔑 <b>Your Access Code:</b>\n"
                f"  <code>{access_code}</code>\n"
                f"  (Use this to login in the mini app)\n\n"
            )
        
        msg += (
            f"🌐 <b>Direct URL:</b>\n"
            f"  <a href=\"{MINI_APP_URL}\">{MINI_APP_URL}</a>\n\n"
            "👆 <b>Tap the button below to open the app!</b>"
        )
        
        try:
            from telegram import InlineKeyboardButton, InlineKeyboardMarkup, WebAppInfo
            
            keyboard_buttons = [
                [InlineKeyboardButton(
                    "📱 Open InternHub Pro",
                    web_app=WebAppInfo(url=MINI_APP_URL)
                )],
            ]
            
            # Add direct link as fallback
            keyboard_buttons.append([
                InlineKeyboardButton(
                    "🌐 Open in Browser",
                    url=MINI_APP_URL
                ),
            ])
            
            keyboard = InlineKeyboardMarkup(keyboard_buttons)
            
            await update.message.reply_text(
                msg,
                parse_mode='HTML',
                reply_markup=keyboard,
                disable_web_page_preview=True,
            )
        except Exception as e:
            logger.debug(f"[{AGENT_ID}] Mini app keyboard failed: {e}")
            # Fallback: send without keyboard
            await update.message.reply_text(
                msg, parse_mode='HTML', disable_web_page_preview=True
            )

    # ================================================================
    # ADMIN SECURITY COMMANDS (admin-only, in admin chat)
    # ================================================================

    @command_error_boundary
    async def _cmd_adduser(self, update, context):
        """Add a new authorized user (admin only).
        Usage: /adduser <username> <telegram_id>
        Example: /adduser johndoe 987654321
        """
        telegram_id = update.effective_user.id if update.effective_user else 0
        if not self.security.is_admin(telegram_id):
            await update.message.reply_text("🔒 Admin-only command.")
            return

        if len(context.args) < 2:
            await update.message.reply_text(
                "Usage: /adduser <username> <telegram_id>\n"
                "Example: /adduser johndoe 987654321\n\n"
                "The user will receive a one-time access code."
            )
            return

        username = context.args[0].lstrip('@').lower()
        try:
            new_user_id = int(context.args[1])
        except ValueError:
            await update.message.reply_text("❌ Invalid Telegram ID. Must be a number.")
            return

        success, code, msg_text = self.security.add_user(username, new_user_id, added_by=telegram_id)

        if success:
            await update.message.reply_text(
                f"✅ <b>User Added Successfully</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"👤 Username: @{username}\n"
                f"🆔 Telegram ID: <code>{new_user_id}</code>\n"
                f"🔑 Access Code: <code>{code}</code>\n\n"
                f"⚠️ Send this code to the user privately.\n"
                f"They need it for the mini-app login.\n"
                f"Code is fully random — no username/ID derivation.",
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(f"⚠️ {msg_text}")

    @command_error_boundary
    async def _cmd_removeuser(self, update, context):
        """Remove/deactivate a user (admin only).
        Usage: /removeuser <telegram_id>
        """
        telegram_id = update.effective_user.id if update.effective_user else 0
        if not self.security.is_admin(telegram_id):
            await update.message.reply_text("🔒 Admin-only command.")
            return

        if not context.args:
            await update.message.reply_text(
                "Usage: /removeuser <telegram_id>\n"
                "Use /listusers to see all users and their IDs."
            )
            return

        try:
            target_id = int(context.args[0])
        except ValueError:
            await update.message.reply_text("❌ Invalid Telegram ID.")
            return

        success, msg_text = self.security.remove_user(target_id)
        emoji = "✅" if success else "⚠️"
        await update.message.reply_text(f"{emoji} {msg_text}", parse_mode='HTML')

    @command_error_boundary
    async def _cmd_readduser(self, update, context):
        """Re-enable a previously removed user (admin only).
        Usage: /readduser <telegram_id>
        """
        telegram_id = update.effective_user.id if update.effective_user else 0
        if not self.security.is_admin(telegram_id):
            await update.message.reply_text("🔒 Admin-only command.")
            return

        if not context.args:
            await update.message.reply_text("Usage: /readduser <telegram_id>")
            return

        try:
            target_id = int(context.args[0])
        except ValueError:
            await update.message.reply_text("❌ Invalid Telegram ID.")
            return

        success, code, msg_text = self.security.readd_user(target_id)
        if success:
            await update.message.reply_text(
                f"✅ <b>User Re-activated</b>\n\n"
                f"{msg_text}\n"
                f"🔑 New Access Code: <code>{code}</code>",
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(f"⚠️ {msg_text}")

    @command_error_boundary
    async def _cmd_listusers(self, update, context):
        """List all authorized users (admin only)."""
        telegram_id = update.effective_user.id if update.effective_user else 0
        if not self.security.is_admin(telegram_id):
            await update.message.reply_text("🔒 Admin-only command.")
            return

        msg = self.security.format_user_list()
        await self._send_long_message(update, msg)

    @command_error_boundary
    async def _cmd_gencode(self, update, context):
        """Regenerate access code for a user (admin only).
        Usage: /gencode <telegram_id>
        """
        telegram_id = update.effective_user.id if update.effective_user else 0
        if not self.security.is_admin(telegram_id):
            await update.message.reply_text("🔒 Admin-only command.")
            return

        if not context.args:
            await update.message.reply_text(
                "Usage: /gencode <telegram_id>\n"
                "Generates a new access code (old code invalidated)."
            )
            return

        try:
            target_id = int(context.args[0])
        except ValueError:
            await update.message.reply_text("❌ Invalid Telegram ID.")
            return

        success, code, msg_text = self.security.regenerate_code(target_id)
        if success:
            await update.message.reply_text(
                f"🔑 <b>New Access Code Generated</b>\n\n"
                f"{msg_text}\n"
                f"New Code: <code>{code}</code>\n\n"
                f"⚠️ Previous code & sessions invalidated.",
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(f"⚠️ {msg_text}")

    @command_error_boundary
    async def _cmd_secstatus(self, update, context):
        """Security dashboard (admin only)."""
        telegram_id = update.effective_user.id if update.effective_user else 0
        if not self.security.is_admin(telegram_id):
            await update.message.reply_text("🔒 Admin-only command.")
            return

        msg = self.security.format_security_dashboard()
        await self._send_long_message(update, msg)

    # ================================================================
    # SCHEDULED REPORTS
    # ================================================================

    async def send_morning_brief(self):
        """Send scheduled morning brief at 07:15 IST."""
        logger.info(f"[{AGENT_ID}] Sending morning brief...")
        data = self.db.get_morning_brief_data()
        # Add source counts for the brief
        try:
            data['source_counts'] = self.db.get_source_counts()
        except Exception:
            data['source_counts'] = {}
        msg = self.formatter.morning_brief(data)
        await self.send_message(msg)

    async def send_evening_summary(self):
        """Send scheduled evening summary at 10:00 PM IST."""
        logger.info(f"[{AGENT_ID}] Sending evening summary...")
        data = {
            'today_total': self.db.count_clean_listings(hours=24),
            'afternoon_new': self.db.count_clean_listings(hours=12),
            'applied_today': self.db.count_outcomes_today(),
            'dark_finds': self.db.count_dark_listings_today(),
        }
        msg = self.formatter.evening_summary(data)
        await self.send_message(msg)

    async def send_urgent_alert(self, company: str, signal_score: float,
                                 signal_type: str):
        """Send urgent alert for high-value signals."""
        msg = (
            f"🚨 <b>URGENT SIGNAL ALERT</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"🏢 <b>{company}</b>\n"
            f"📡 Signal: {signal_type}\n"
            f"💪 Score: {signal_score:.0f}/100\n\n"
            f"💡 Use /research {company} for details"
        )
        await self.send_message(msg)

    async def send_blue_ocean_alert(self, listing: Dict):
        """Send Blue Ocean discovery alert."""
        title = listing.get('title', 'Unknown')
        company = listing.get('company', 'Unknown')
        applicants = listing.get('applicants', 0)
        lid = listing.get('id', 0)

        msg = (
            f"🌊 <b>BLUE OCEAN ALERT!</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"<b>{title}</b> @ {company}\n"
            f"👥 Only {applicants} applicants!\n\n"
            f"Quick: /package {lid}"
        )
        await self.send_message(msg)

    # ================================================================
    # UTILITY METHODS
    # ================================================================

    async def _send_long_message(self, update, text: str):
        """Send a message, splitting if too long. Smart split on newlines."""
        if len(text) <= TG_MAX_LEN:
            try:
                await update.message.reply_text(text, parse_mode='HTML')
            except Exception:
                try:
                    await update.message.reply_text(text)
                except Exception as e:
                    logger.error(f"[{AGENT_ID}] Message send error: {e}")
            return

        # Smart split: try to break on newlines to avoid cutting HTML tags
        chunks = []
        remaining = text
        while remaining:
            if len(remaining) <= TG_MAX_LEN:
                chunks.append(remaining)
                break

            # Find the last newline within the limit
            split_pos = remaining.rfind('\n', 0, TG_MAX_LEN)
            if split_pos == -1 or split_pos < TG_MAX_LEN // 2:
                # No good newline split point, hard split at limit
                split_pos = TG_MAX_LEN

            chunks.append(remaining[:split_pos])
            remaining = remaining[split_pos:].lstrip('\n')

        for chunk in chunks:
            if not chunk.strip():
                continue
            try:
                await update.message.reply_text(chunk, parse_mode='HTML')
            except Exception:
                try:
                    await update.message.reply_text(chunk)
                except Exception as e:
                    logger.error(f"[{AGENT_ID}] Message send error: {e}")
                    break

    # ================================================================
    # BOT MENU SETUP — Persistent keyboard + Open App button
    # ================================================================

    async def _setup_bot_menu(self):
        """
        Set up the Telegram bot menu:
        1. Register /commands with BotFather via setMyCommands
        2. Set MenuButtonWebApp if MINI_APP_URL is configured
           (shows 'Open App' button at bottom-left like screenshot 3)
        """
        if not self._app or not self._app.bot:
            return

        try:
            from telegram import BotCommand, MenuButtonWebApp, WebAppInfo, MenuButtonCommands
            
            # Step 1: Set bot commands (shows in / menu)
            bot_commands = [
                BotCommand("jobs", "📋 Browse filtered internships"),
                BotCommand("morning", "🌅 Morning brief report"),
                BotCommand("top", "🏆 Top listings by PPO score"),
                BotCommand("ocean", "🌊 Blue Ocean opportunities"),
                BotCommand("run", "🚀 Run agents/pipeline"),
                BotCommand("health", "💚 System health check"),
                BotCommand("stats", "📈 Weekly statistics"),
                BotCommand("export", "📤 Export to Excel"),
                BotCommand("miniapp", "📱 Open InternHub Pro"),
                BotCommand("menu", "📋 Show command menu"),
                BotCommand("help", "📖 Full command reference"),
                BotCommand("startpipeline", "▶️ Start full pipeline (admin)"),
                BotCommand("stoppipeline", "⏹ Stop pipeline (admin)"),
                BotCommand("secstatus", "🔐 Security dashboard (admin)"),
            ]
            await self._app.bot.set_my_commands(bot_commands)
            logger.info(f"[{AGENT_ID}] Bot commands registered ({len(bot_commands)} commands)")
            
            # Step 2: Set Mini App button if configured
            if MINI_APP_URL:
                try:
                    menu_button = MenuButtonWebApp(
                        text="Open App",
                        web_app=WebAppInfo(url=MINI_APP_URL)
                    )
                    await self._app.bot.set_chat_menu_button(menu_button=menu_button)
                    logger.info(f"[{AGENT_ID}] Mini App menu button set: {MINI_APP_URL}")
                except Exception as e:
                    logger.warning(f"[{AGENT_ID}] Failed to set MenuButtonWebApp: {e}")
                    # Fallback to regular commands menu
                    try:
                        await self._app.bot.set_chat_menu_button(
                            menu_button=MenuButtonCommands()
                        )
                    except Exception:
                        pass
            else:
                # Set standard commands menu button
                try:
                    await self._app.bot.set_chat_menu_button(
                        menu_button=MenuButtonCommands()
                    )
                except Exception:
                    pass
                    
        except ImportError as e:
            logger.warning(f"[{AGENT_ID}] Menu setup import error: {e}")
        except Exception as e:
            logger.warning(f"[{AGENT_ID}] Menu setup error (non-fatal): {e}")

    def _get_main_reply_keyboard(self, is_admin: bool = False):
        """
        Build a persistent ReplyKeyboardMarkup with organized command buttons.
        Layout matches the 3-column grid style from the reference screenshot.
        """
        from telegram import ReplyKeyboardMarkup, KeyboardButton, WebAppInfo
        
        # Row 1: Core browsing
        row1 = [
            KeyboardButton("📋 Jobs"),
            KeyboardButton("🏆 Top 20"),
            KeyboardButton("🌊 Ocean"),
        ]
        
        # Row 2: Reports
        row2 = [
            KeyboardButton("🌅 Morning"),
            KeyboardButton("📈 Stats"),
            KeyboardButton("💚 Health"),
        ]
        
        # Row 3: Actions
        row3 = [
            KeyboardButton("🚀 Run Pipeline"),
            KeyboardButton("📤 Export"),
            KeyboardButton("🔍 Sources"),
        ]
        
        # Row 4: Tools
        row4 = [
            KeyboardButton("📂 Browse"),
            KeyboardButton("💰 Quota"),
            KeyboardButton("📖 Help"),
        ]
        
        rows = [row1, row2, row3, row4]
        
        # Row 5: Mini App + Admin (if applicable)
        if MINI_APP_URL:
            row5 = [KeyboardButton("📱 Mini App")]
        else:
            row5 = []
            
        if is_admin:
            row5.extend([
                KeyboardButton("🔐 Security"),
                KeyboardButton("⏹ Stop Pipeline"),
            ])
        
        if row5:
            rows.append(row5)
        
        return ReplyKeyboardMarkup(
            rows,
            resize_keyboard=True,
            is_persistent=True,
            input_field_placeholder="Type a command or tap a button..."
        )

    async def _handle_menu_button_press(self, update, context):
        """
        Handle text messages from the persistent reply keyboard buttons.
        Maps button text to the corresponding command handler.
        """
        if not update.message or not update.message.text:
            return
        
        text = update.message.text.strip()
        
        # Map button labels to command handlers
        button_map = {
            "📋 Jobs": self._cmd_jobs,
            "🏆 Top 20": self._cmd_top,
            "🌊 Ocean": self._cmd_ocean,
            "🌅 Morning": self._cmd_morning,
            "📈 Stats": self._cmd_stats,
            "💚 Health": self._cmd_health,
            "🚀 Run Pipeline": self._cmd_startpipeline,
            "📤 Export": self._cmd_export,
            "🔍 Sources": self._cmd_sources,
            "📂 Browse": self._cmd_browse,
            "💰 Quota": self._cmd_quota,
            "📖 Help": self._cmd_help,
            "📱 Mini App": self._cmd_miniapp,
            "🔐 Security": self._cmd_secstatus,
            "⏹ Stop Pipeline": self._cmd_stoppipeline,
        }
        
        handler = button_map.get(text)
        if handler:
            # Set default args for commands that need them
            if text == "🏆 Top 20":
                context.args = ['20']
            elif text in ("🚀 Run Pipeline",):
                context.args = ['pipeline']
            else:
                context.args = []
            await handler(update, context)

    @command_error_boundary
    async def _cmd_menu(self, update, context):
        """Send the persistent reply keyboard menu."""
        telegram_id = update.effective_user.id if update.effective_user else 0
        is_admin = self.security.is_admin(telegram_id)
        
        keyboard = self._get_main_reply_keyboard(is_admin=is_admin)
        await update.message.reply_text(
            "📋 <b>Command Menu</b>\n"
            "Tap any button below to execute a command.\n"
            "You can also type /commands directly.\n\n"
            "💡 The menu stays at the bottom for quick access.",
            parse_mode='HTML',
            reply_markup=keyboard,
        )

    # ================================================================
    # ADMIN PIPELINE COMMANDS — start/stop pipeline (admin only)
    # ================================================================

    @command_error_boundary
    async def _cmd_startpipeline(self, update, context):
        """Start the full pipeline (admin only).
        Usage: /startpipeline
        Alias: 🚀 Run Pipeline button
        """
        telegram_id = update.effective_user.id if update.effective_user else 0
        if not self.security.is_admin(telegram_id):
            await update.message.reply_text("🔒 Pipeline control is admin-only.")
            return

        self._pipeline_stop_requested = False
        # Delegate to /run pipeline
        context.args = ['pipeline']
        await self._cmd_run(update, context)

    @command_error_boundary
    async def _cmd_stoppipeline(self, update, context):
        """Stop a running pipeline (admin only).
        Usage: /stoppipeline
        Cancels all running pipeline tasks.
        """
        telegram_id = update.effective_user.id if update.effective_user else 0
        if not self.security.is_admin(telegram_id):
            await update.message.reply_text("🔒 Pipeline control is admin-only.")
            return

        self._pipeline_stop_requested = True
        stopped = []
        
        async with self._task_lock:
            for name, info in list(self._running_tasks.items()):
                task = info.get('task')
                if task and not task.done():
                    task.cancel()
                    stopped.append(name)
        
        if stopped:
            await update.message.reply_text(
                f"⏹ <b>Pipeline Stop Requested</b>\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"Stopping: {', '.join(stopped)}\n"
                f"Tasks will finish their current operation and stop.\n\n"
                f"Use /status to verify all tasks stopped.",
                parse_mode='HTML'
            )
        else:
            await update.message.reply_text(
                "💤 No pipeline tasks are currently running.\n"
                "Use /startpipeline or /run pipeline to start."
            )

    # ================================================================
    # UTILITY METHODS
    # ================================================================

    @staticmethod
    def _parse_listing_id(args) -> Optional[int]:
        """Parse listing ID from command arguments."""
        if not args:
            return None
        try:
            return int(args[0])
        except (ValueError, IndexError):
            return None


# ============================================================
# SINGLETON ACCESS
# ============================================================

_reporter_instance: Optional[TelegramReporter] = None


def get_telegram_reporter() -> TelegramReporter:
    """Get or create the singleton TelegramReporter instance."""
    global _reporter_instance
    if _reporter_instance is None:
        _reporter_instance = TelegramReporter()
    return _reporter_instance


# ============================================================
# SELF-TEST
# ============================================================

if __name__ == "__main__":
    print("=" * 60)
    print(f"  {AGENT_NAME} ({AGENT_ID}) — Self-Test")
    print("=" * 60)

    # Test ReportFormatter
    formatter = ReportFormatter()

    # Test morning brief format
    test_data = {
        'total_new': 150,
        'after_ghost_filter': 95,
        'blue_ocean_count': 3,
        'signals_fired': 7,
        'top_10': [
            {'title': 'Marketing Intern', 'company': 'McKinsey', 'ppo_score': 85.2,
             'is_blue_ocean': True, 'is_ppo': True, 'stipend_monthly': 50000},
        ],
        'dark_finds': [],
        'urgent_deadlines': [],
    }
    print("\nMorning Brief Preview:")
    print(formatter.morning_brief(test_data)[:500])

    print(f"\nCommands registered: 26")
    print(f"Message max length: {TG_MAX_LEN}")
    print(f"Valid outcomes: {', '.join(VALID_OUTCOMES)}")
    print(f"\n✅ {AGENT_NAME} ({AGENT_ID}) ready!")
    print("=" * 60)
